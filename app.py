from __future__ import annotations

import csv
import hashlib
import io
import zipfile
import os
import re
import sqlite3
import time
import traceback
import unicodedata
import zipfile
from collections import defaultdict
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import Element, SubElement, tostring

import pandas as pd
from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from werkzeug.utils import secure_filename
from jinja2 import DictLoader

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("DATABASE_PATH", BASE_DIR / "fleetflow.db"))
UPLOAD_DIR = BASE_DIR / "uploads"
REPORT_DIR = BASE_DIR / "reports"
SEED_XLSX = BASE_DIR / "BANCO_DE_DADOS.xlsx"
if not SEED_XLSX.exists():
    SEED_XLSX = BASE_DIR / "data" / "BANCO_DE_DADOS.xlsx"
UPLOAD_DIR.mkdir(exist_ok=True)
REPORT_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.jinja_loader = DictLoader({'closing_detail.html': '{% extends \'base.html\' %}{% block title %}{{closing.label}}{% endblock %}{% block subtitle %}{{closing.status}} · Confira antes de enviar o XML ao banco.{% endblock %}{% block content %}<div class="actions"><a class="button" href="{{url_for(\'closing_excel\',closing_id=closing.id)}}">Baixar Excel</a><a class="button" href="{{url_for(\'closing_xml\',closing_id=closing.id)}}">Baixar XML SEPA</a></div><section class="panel table-wrap"><table><thead><tr><th>Motorista</th><th>Origem</th><th>IBAN</th><th>Bruto</th><th>Dinheiro</th><th>Comissão</th><th>Combustível</th><th>Desconto</th><th>Reembolso</th><th>Taxa</th><th>Líquido</th><th>Total IBAN</th><th>Recibo</th></tr></thead><tbody>{% for i in items %}<tr class="{{\'problem\' if not i.iban or i.net_before_group<0}}"><td>{{i.driver_name}}</td><td>{{i.origins}}</td><td>{{i.iban or \'SEM IBAN\'}}</td><td>€ {{\'%.2f\'|format(i.gross)}}</td><td>€ {{\'%.2f\'|format(i.cash)}}</td><td>€ {{\'%.2f\'|format(i.commission)}}</td><td>€ {{\'%.2f\'|format(i.fuel)}}</td><td>€ {{\'%.2f\'|format(i.discount)}}</td><td>€ {{\'%.2f\'|format(i.reimbursement)}}</td><td>€ {{\'%.2f\'|format(i.bank_fee)}}</td><td>€ {{\'%.2f\'|format(i.net_before_group-i.bank_fee)}}</td><td><b>€ {{\'%.2f\'|format(i.group_total)}}</b></td></tr>{% endfor %}</tbody></table></section>{% endblock %}\n', 'base.html': '<!doctype html><html lang="pt"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{{ company_name }} · FleetFlow</title><script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script><style>\n:root{--bg:#f4f7fb;--card:#fff;--text:#172033;--muted:#718096;--accent:#315efb;--dark:#101828;--border:#e5e9f2;--danger:#d92d20}*{box-sizing:border-box}body{margin:0;font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--bg);color:var(--text)}.layout{display:grid;grid-template-columns:235px 1fr;min-height:100vh}aside{background:var(--dark);color:#fff;padding:24px 16px}.brand{display:flex;align-items:center;gap:12px;margin:0 8px 28px}.brand>span,.logo-big{display:grid;place-items:center;background:var(--accent);color:#fff;border-radius:13px;font-weight:800}.brand>span{width:42px;height:42px}.brand small{display:block;color:#98a2b3;margin-top:3px}nav{display:grid;gap:6px}nav a{color:#d0d5dd;text-decoration:none;padding:12px 14px;border-radius:10px}nav a:hover{background:#1d2939;color:#fff}main{padding:30px;min-width:0}header h1{margin:0;font-size:28px}header p{margin:6px 0 24px;color:var(--muted)}.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin-bottom:18px}.metric,.panel{background:var(--card);border:1px solid var(--border);border-radius:16px}.metric{padding:18px}.metric span{display:block;color:var(--muted);font-size:13px}.metric b{display:block;margin-top:8px;font-size:24px}.metric.alert b{color:var(--danger)}.panel{padding:20px;margin-bottom:18px}.panel h2{margin:0 0 14px;font-size:18px}.grid2{display:grid;grid-template-columns:1fr 1fr;gap:18px}.toolbar,.actions{display:flex;gap:10px;align-items:center;margin-bottom:18px}.toolbar input{flex:1}input,select,textarea{width:100%;border:1px solid #d0d5dd;border-radius:10px;padding:11px 12px;font:inherit;background:#fff}textarea{min-height:100px}button,.button{border:0;border-radius:10px;background:var(--accent);color:#fff;padding:11px 16px;font-weight:700;text-decoration:none;cursor:pointer;display:inline-block}.secondary{color:var(--accent);text-decoration:none;padding:10px}.danger{background:var(--danger)}label{display:grid;gap:7px;font-size:13px;font-weight:600}.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}.full{grid-column:1/-1}.table-wrap{overflow:auto}table{width:100%;border-collapse:collapse;font-size:13px}th,td{text-align:left;padding:11px 10px;border-bottom:1px solid var(--border);white-space:nowrap}th{color:var(--muted);font-size:12px;text-transform:uppercase}.link{color:var(--accent);font-weight:700;text-decoration:none}.badge{display:inline-block;border-radius:999px;padding:4px 8px;font-size:11px;font-weight:800;background:#edf2f7}.badge.azul{background:#dbeafe;color:#1d4ed8}.badge.yellow{background:#fef3c7;color:#92400e}.flash{padding:12px 14px;border-radius:10px;margin-bottom:16px;background:#eaf2ff}.flash.danger{background:#fee4e2;color:#b42318}.flash.warning{background:#fef0c7;color:#93370d}.flash.success{background:#dcfae6;color:#067647}.note{color:var(--muted)}tr.problem{background:#fff4ed}.login-body{min-height:100vh;display:grid;place-items:center;background:#101828}.login-card{width:min(410px,calc(100% - 32px));background:#fff;border-radius:20px;padding:32px;display:grid;gap:16px}.login-card h1,.login-card p{margin:0;text-align:center}.login-card p,.login-card small{color:var(--muted);text-align:center}.logo-big{width:58px;height:58px;margin:auto;font-size:20px}@media(max-width:950px){.layout{grid-template-columns:1fr}aside{padding:14px}.brand{margin-bottom:12px}nav{display:flex;overflow:auto}.metrics{grid-template-columns:repeat(2,1fr)}.grid2,.form-grid{grid-template-columns:1fr}main{padding:18px}}@media(max-width:520px){.metrics{grid-template-columns:1fr}.toolbar,.actions{align-items:stretch;flex-direction:column}.toolbar input,.toolbar button,.actions a{width:100%}}\n\n</style></head><body>\n<div class="layout"><aside><div class="brand"><span>FF</span><div><b>FleetFlow</b><small>{{ company_name }} · v{{ app_version }}</small></div></div><nav><a href="{{ url_for(\'dashboard\') }}">Dashboard</a><a href="{{ url_for(\'drivers\') }}">Motoristas</a><a href="{{ url_for(\'database_page\') }}">Banco de dados</a><a href="{{ url_for(\'imports_page\') }}">Importações</a><a href="{{ url_for(\'closings_page\') }}">Fechamentos</a><a href="{{ url_for(\'missing_iban_page\') }}">Sem IBAN</a><a href="{{ url_for(\'partners_page\') }}">Parceiros</a><a href="{{ url_for(\'logout\') }}">Sair</a></nav></aside><main><header><div><h1>{% block title %}{% endblock %}</h1><p>{% block subtitle %}{% endblock %}</p></div></header>{% for cat,msg in get_flashed_messages(with_categories=true) %}<div class="flash {{cat}}">{{msg}}</div>{% endfor %}{% block content %}{% endblock %}</main></div></body></html>\n', 'closings.html': '{% extends \'base.html\' %}{% block title %}Fechamentos{% endblock %}{% block subtitle %}Processe depois de atualizar o banco e importar todos os relatórios da semana.{% endblock %}{% block content %}<section class="panel"><h2>Novo fechamento</h2><form method="post" class="toolbar"><input name="label" placeholder="Ex.: Semana 20/07/2026"><button>Processar semana</button></form><p class="note">A taxa de € 1,25 é aplicada uma única vez por IBAN AZUL, após a consolidação.</p></section><section class="panel"><table><thead><tr><th>ID</th><th>Nome</th><th>Data</th><th>Status</th><th></th></tr></thead><tbody>{% for r in rows %}<tr><td>#{{r.id}}</td><td>{{r.label}}</td><td>{{r.created_at}}</td><td>{{r.status}}</td><td><a class="link" href="{{url_for(\'closing_detail\',closing_id=r.id)}}">Conferir</a></td></tr>{% else %}<tr><td colspan="5">Nenhum fechamento.</td></tr>{% endfor %}</tbody></table></section>{% endblock %}\n', 'login.html': '<!doctype html><html lang="pt"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Entrar · FleetFlow</title><link rel="stylesheet" href="{{ url_for(\'static\', filename=\'style.css\') }}"></head><body class="login-body"><form class="login-card" method="post"><div class="logo-big">FF</div><h1>FleetFlow</h1><p>{{ company_name }}</p>{% for cat,msg in get_flashed_messages(with_categories=true) %}<div class="flash {{cat}}">{{msg}}</div>{% endfor %}<label>Utilizador<input name="username" required autofocus></label><label>Senha<input name="password" type="password" required></label><button>Entrar</button><small>Configure ADMIN_USER e ADMIN_PASSWORD no Render.</small></form></body></html>\n', 'database.html': '{% extends \'base.html\' %}{% block title %}Banco de dados{% endblock %}{% block subtitle %}Base online dos motoristas e mapa Arquivos_Pgto.{% endblock %}{% block content %}<div class="metrics"><div class="metric"><span>Motoristas</span><b>{{n}}</b></div><div class="metric"><span>Mapas de leitura</span><b>{{m}}</b></div></div><section class="panel"><h2>Atualizar banco completo</h2><p>Envie um XLSX com as abas <b>Banco_de_Dados</b> e <b>Arquivos_Pgto</b>. A importação substitui o cadastro atual.</p><form method="post" enctype="multipart/form-data"><input type="file" name="database_file" accept=".xlsx" required><button>Importar banco atualizado</button></form></section><section class="panel"><h2>Alterações rápidas</h2><p>Para alterar somente IBAN, banco AZUL/YELLOW, porcentagem, desconto, parceiro ou cartão, use a página <a class="link" href="{{url_for(\'drivers\')}}">Motoristas</a>.</p></section>{% endblock %}\n', 'drivers.html': '{% extends \'base.html\' %}{% block title %}Motoristas{% endblock %}{% block subtitle %}{{total}} cadastros. A tabela mostra até 500 resultados por pesquisa.{% endblock %}{% block content %}<form class="toolbar"><input name="q" value="{{q}}" placeholder="Nome, ID, IBAN, cidade ou parceiro"><button>Pesquisar</button></form><section class="panel table-wrap"><table><thead><tr><th>Motorista</th><th>ID</th><th>Cidade</th><th>Companhia</th><th>IBAN</th><th>Banco</th><th>%</th><th>Parceiro</th><th></th></tr></thead><tbody>{% for d in rows %}<tr><td><b>{{d.name}}</b></td><td>{{d.external_id}}</td><td>{{d.city}}</td><td>{{d.company}}</td><td>{{d.iban}}</td><td><span class="badge {{d.bank_color|lower}}">{{d.bank_color}}</span></td><td>{{\'%.2f\'|format(d.percentage*100)}}%</td><td>{{d.partner}}</td><td><a class="link" href="{{url_for(\'edit_driver\',driver_id=d.id)}}">Editar</a></td></tr>{% endfor %}</tbody></table></section>{% endblock %}\n', 'imports.html': '{% extends \'base.html\' %}{% block title %}Importações da semana{% endblock %}{% block subtitle %}Envie vários CSVs de uma vez e, separadamente, os ficheiros PRIO em XLSX.{% endblock %}{% block content %}<div class="grid2"><section class="panel"><h2>Relatórios das plataformas</h2><form method="post" enctype="multipart/form-data"><input type="hidden" name="kind" value="reports"><input type="file" name="files" accept=".csv" multiple required><button>Importar relatórios</button></form></section><section class="panel"><h2>Cartões combustível</h2><form method="post" enctype="multipart/form-data"><input type="hidden" name="kind" value="fuel"><input type="file" name="files" accept=".xlsx,.xls" multiple required><button>Importar PRIO</button></form></section></div><form method="post" action="{{url_for(\'clear_imports\')}}" onsubmit="return confirm(\'Limpar todos os ficheiros da semana?\')"><button class="danger">Limpar semana importada</button></form><section class="panel table-wrap"><h2>Histórico</h2><table><thead><tr><th>Ficheiro</th><th>Tipo</th><th>Status</th><th>Linhas</th><th>Mensagem</th><th>Data</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.filename}}</td><td>{{r.kind}}</td><td><span class="badge {{\'azul\' if r.status==\'OK\' else \'yellow\'}}">{{r.status}}</span></td><td>{{r.rows_count}}</td><td>{{r.message}}</td><td>{{r.created_at}}</td></tr>{% else %}<tr><td colspan="6">Nenhum ficheiro importado.</td></tr>{% endfor %}</tbody></table></section>{% endblock %}\n', 'error.html': '{% extends \'base.html\' %}{% block title %}Erro no sistema{% endblock %}{% block subtitle %}Não foi possível concluir esta operação.{% endblock %}{% block content %}<section class="panel"><h2>Erro interno</h2><p>{{ message }}</p><p><a class="link" href="{{ url_for(\'login\') }}">Voltar ao login</a></p></section>{% endblock %}', 'dashboard.html': '{% extends \'base.html\' %}{% block title %}Dashboard semanal{% endblock %}{% block subtitle %}Visão geral do último fechamento processado.{% endblock %}{% block content %}\n<div class="metrics"><div class="metric"><span>Motoristas</span><b>{{drivers}}</b></div><div class="metric"><span>Ficheiros importados</span><b>{{imports}}</b></div><div class="metric"><span>Bruto</span><b>€ {{\'%.2f\'|format(stats.gross)}}</b></div><div class="metric"><span>Líquido</span><b>€ {{\'%.2f\'|format(stats.net)}}</b></div><div class="metric"><span>Comissão</span><b>€ {{\'%.2f\'|format(stats.commission)}}</b></div><div class="metric"><span>Combustível</span><b>€ {{\'%.2f\'|format(stats.fuel)}}</b></div><div class="metric"><span>Transferências</span><b>{{stats.payments}}</b></div><div class="metric alert"><span>Sem IBAN</span><b>{{stats.missing_iban}}</b></div></div>\n<div class="grid2"><section class="panel"><h2>Ganhos por origem</h2><canvas id="originChart"></canvas></section><section class="panel"><h2>Distribuição financeira</h2><canvas id="moneyChart"></canvas></section></div>\n<section class="panel"><h2>Fechamentos recentes</h2><table><thead><tr><th>ID</th><th>Semana</th><th>Data</th><th>Status</th><th></th></tr></thead><tbody>{% for r in recent %}<tr><td>#{{r.id}}</td><td>{{r.label}}</td><td>{{r.created_at}}</td><td>{{r.status}}</td><td><a class="link" href="{{url_for(\'closing_detail\',closing_id=r.id)}}">Abrir</a></td></tr>{% else %}<tr><td colspan="5">Nenhum fechamento processado.</td></tr>{% endfor %}</tbody></table></section>\n<script>new Chart(document.getElementById(\'originChart\'),{type:\'bar\',data:{labels:{{ origin_labels|tojson }},datasets:[{label:\'Bruto (€)\',data:{{ origin_values|tojson }}}]},options:{responsive:true}});new Chart(document.getElementById(\'moneyChart\'),{type:\'doughnut\',data:{labels:[\'Comissão\',\'Combustível\',\'Descontos\',\'Reembolsos\'],datasets:[{data:[{{stats.commission}},{{stats.fuel}},{{stats.discount}},{{stats.reimbursement}}]}]},options:{responsive:true}});</script>\n{% endblock %}\n', 'driver_edit.html': '{% extends \'base.html\' %}{% block title %}Editar motorista{% endblock %}{% block subtitle %}As alterações passam a valer no próximo processamento.{% endblock %}{% block content %}<form method="post" class="panel form-grid"><label>Nome<input name="name" value="{{d.name}}" required></label><label>ID / e-mail / telefone<input name="external_id" value="{{d.external_id}}"></label><label>Cidade/campanha<input name="city" value="{{d.city}}"></label><label>Companhia<input name="company" value="{{d.company}}"></label><label>IBAN<input name="iban" value="{{d.iban}}"></label><label>Banco<select form="save-{{d.id}}" name="bank_color"><option {{\'selected\' if d.bank_color==\'YELLOW\'}}>YELLOW</option><option {{\'selected\' if d.bank_color==\'AZUL\'}}>AZUL</option></select></label><label>Porcentagem (decimal)<input name="percentage" type="number" step="0.0001" value="{{d.percentage}}"></label><label>Comissão Muryllo<input name="commission_owner" type="number" step="0.01" value="{{d.commission_owner}}"></label><label>Parceiro<input name="partner" value="{{d.partner}}"></label><label>Comissão parceiro<input name="partner_commission" type="number" step="0.01" value="{{d.partner_commission}}"></label><label>Cartão combustível<input name="fuel_card" value="{{d.fuel_card}}"></label><label>Aluguel/categoria<input name="rental_label" value="{{d.rental_label}}"></label><label>Desconto<input name="discount" type="number" step="0.01" value="{{d.discount}}"></label><label>Reembolso<input name="reimbursement" type="number" step="0.01" value="{{d.reimbursement}}"></label><label>Imediata<input name="immediate" type="number" step="0.01" value="{{d.immediate}}"></label><label class="full">Observação<textarea name="observation">{{d.observation}}</textarea></label><div class="full actions"><button>Guardar alterações</button><a class="secondary" href="{{url_for(\'drivers\')}}">Voltar</a></div></form>{% endblock %}\n'})

app.jinja_loader.mapping["missing_iban.html"] = """{% extends 'base.html' %}{% block title %}Sem IBAN{% endblock %}{% block subtitle %}Corrija os dados bancários dos motoristas pendentes e atualize o último fechamento sem refazer as importações.{% endblock %}{% block content %}<div class="metrics"><div class="metric alert"><span>Pendentes no último fechamento</span><b>{{closing_missing}}</b></div><div class="metric"><span>Cadastros sem IBAN</span><b>{{total_missing}}</b></div></div><form class="toolbar"><input name="q" value="{{q}}" placeholder="Pesquisar nome, ID, parceiro ou cidade"><button>Pesquisar</button><a class="button" href="{{url_for('missing_iban_excel')}}">Baixar Excel</a></form><section class="panel table-wrap"><table><thead><tr><th>Motorista</th><th>ID</th><th>Cidade</th><th>Companhia</th><th>Parceiro</th><th>IBAN</th><th>Banco</th><th>Comissão parceiro</th><th>Líquido pendente</th><th></th></tr></thead><tbody>{% for d in rows %}<tr class="problem"><td><form id="save-{{d.id}}" method="post" action="{{url_for('save_missing_iban',driver_id=d.id)}}"></form><b>{{d.name}}</b></td><td>{{d.external_id}}</td><td>{{d.city}}</td><td>{{d.company}}</td><td><input form="save-{{d.id}}" name="partner" value="{{d.partner}}" style="min-width:130px"></td><td><input form="save-{{d.id}}" name="iban" value="{{d.iban}}" placeholder="PT50..." required style="min-width:250px"></td><td><select form="save-{{d.id}}" name="bank_color"><option value="YELLOW" {{'selected' if d.bank_color=='YELLOW'}}>YELLOW</option><option value="AZUL" {{'selected' if d.bank_color=='AZUL'}}>AZUL</option></select></td><td><input form="save-{{d.id}}" name="partner_commission" type="number" step="0.01" value="{{d.partner_commission}}" style="min-width:110px"></td><td>€ {{'%.2f'|format(d.pending_net or 0)}}</td><td><button form="save-{{d.id}}">Salvar e recalcular</button></td></tr>{% else %}<tr><td colspan="10">Nenhum motorista sem IBAN encontrado.</td></tr>{% endfor %}</tbody></table></section><p class="note">Depois de salvar, o motorista sai desta lista. Se existir um fechamento processado, o IBAN, a taxa bancária e o total consolidado são atualizados imediatamente.</p>{% endblock %}"""
app.jinja_loader.mapping["partners.html"] = """{% extends 'base.html' %}{% block title %}Parceiros{% endblock %}{% block subtitle %}Relatórios separados por parceiro no último fechamento.{% endblock %}{% block content %}<section class="panel table-wrap"><table><thead><tr><th>Parceiro</th><th>Motoristas</th><th>Bruto</th><th>Líquido</th><th></th></tr></thead><tbody>{% for r in rows %}<tr><td><b>{{r.partner}}</b></td><td>{{r.drivers}}</td><td>€ {{'%.2f'|format(r.gross)}}</td><td>€ {{'%.2f'|format(r.net)}}</td><td><a class="button" href="{{url_for('partner_excel',closing_id=closing_id,partner=r.partner)}}">Baixar Excel</a></td></tr>{% else %}<tr><td colspan="5">Nenhum parceiro neste fechamento.</td></tr>{% endfor %}</tbody></table></section>{% endblock %}"""

app.jinja_loader.mapping["base.html"] = r"""<!doctype html><html lang="pt"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{{ company_name }} · Gestão de Frotas</title><script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script><style>
:root{--bg:#f5f7fb;--card:#fff;--text:#172033;--muted:#667085;--accent:#2563eb;--accent2:#06b6d4;--dark:#0b1220;--border:#e4e7ec;--danger:#d92d20;--success:#079455;--warning:#dc6803;--shadow:0 12px 36px rgba(16,24,40,.08)}*{box-sizing:border-box}body{margin:0;font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:linear-gradient(180deg,#eef3fb 0,#f8fafc 220px);color:var(--text)}.layout{display:grid;grid-template-columns:265px 1fr;min-height:100vh}aside{background:linear-gradient(180deg,#0b1220,#111c31);color:#fff;padding:24px 16px;position:sticky;top:0;height:100vh}.brand{display:flex;align-items:center;gap:12px;margin:0 8px 28px;padding-bottom:20px;border-bottom:1px solid rgba(255,255,255,.1)}.brand-logo,.logo-big{display:grid;place-items:center;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#fff;border-radius:15px;font-weight:900;box-shadow:0 10px 24px rgba(37,99,235,.3)}.brand-logo{width:46px;height:46px;font-size:17px}.brand-img{display:block;border-radius:15px}.brand b{font-size:17px}.brand small{display:block;color:#98a2b3;margin-top:3px}nav{display:grid;gap:5px}nav a{color:#d0d5dd;text-decoration:none;padding:12px 14px;border-radius:11px;font-weight:600;font-size:14px}nav a:hover{background:rgba(255,255,255,.1);color:#fff}.nav-section{padding:12px 14px 5px;color:#667085;font-size:10px;font-weight:800;letter-spacing:.12em;text-transform:uppercase}main{padding:30px;min-width:0;max-width:1700px;width:100%;margin:auto}header{display:flex;justify-content:space-between;align-items:flex-start}header h1{margin:0;font-size:30px;letter-spacing:-.03em}header p{margin:7px 0 24px;color:var(--muted)}.metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px;margin-bottom:18px}.metric,.panel{background:rgba(255,255,255,.96);border:1px solid var(--border);border-radius:18px;box-shadow:var(--shadow)}.metric{padding:18px;position:relative;overflow:hidden}.metric:after{content:"";position:absolute;right:-28px;top:-28px;width:85px;height:85px;border-radius:50%;background:rgba(37,99,235,.07)}.metric span{display:block;color:var(--muted);font-size:13px;font-weight:600}.metric b{display:block;margin-top:8px;font-size:24px;letter-spacing:-.03em}.metric.alert b{color:var(--danger)}.metric.success b{color:var(--success)}.metric.warning b{color:var(--warning)}.panel{padding:20px;margin-bottom:18px}.panel h2{margin:0 0 14px;font-size:18px}.grid2{display:grid;grid-template-columns:1fr 1fr;gap:18px}.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:18px}.toolbar,.actions{display:flex;gap:10px;align-items:center;margin-bottom:18px;flex-wrap:wrap}.toolbar input{flex:1;min-width:220px}input,select,textarea{width:100%;border:1px solid #d0d5dd;border-radius:11px;padding:11px 12px;font:inherit;background:#fff}input:focus,select:focus,textarea:focus{outline:3px solid rgba(37,99,235,.12);border-color:var(--accent)}textarea{min-height:100px}button,.button{border:0;border-radius:11px;background:linear-gradient(135deg,var(--accent),#1d4ed8);color:#fff;padding:11px 16px;font-weight:750;text-decoration:none;cursor:pointer;display:inline-block;box-shadow:0 6px 14px rgba(37,99,235,.2)}.secondary{color:var(--accent);text-decoration:none;padding:10px}.danger{background:var(--danger)}label{display:grid;gap:7px;font-size:13px;font-weight:650}.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}.full{grid-column:1/-1}.table-wrap{overflow:auto}table{width:100%;border-collapse:collapse;font-size:13px}th,td{text-align:left;padding:12px 10px;border-bottom:1px solid var(--border);white-space:nowrap}th{color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.05em}tbody tr:hover{background:#f8fafc}.link{color:var(--accent);font-weight:700;text-decoration:none}.badge{display:inline-block;border-radius:999px;padding:5px 9px;font-size:11px;font-weight:800;background:#edf2f7}.badge.azul{background:#dbeafe;color:#1d4ed8}.badge.yellow{background:#fef3c7;color:#92400e}.flash{padding:12px 14px;border-radius:11px;margin-bottom:16px;background:#eaf2ff}.flash.danger{background:#fee4e2;color:#b42318}.flash.warning{background:#fef0c7;color:#93370d}.flash.success{background:#dcfae6;color:#067647}.note{color:var(--muted)}tr.problem{background:#fff4ed}.quick{display:flex;gap:10px;flex-wrap:wrap}.quick a{background:#fff;border:1px solid var(--border);padding:10px 13px;border-radius:11px;text-decoration:none;color:var(--text);font-weight:700}.quick a:hover{border-color:var(--accent);color:var(--accent)}@media(max-width:1050px){.layout{grid-template-columns:1fr}aside{position:static;height:auto;padding:14px}.brand{margin-bottom:10px;padding-bottom:10px}nav{display:flex;overflow:auto}.nav-section{display:none}.metrics{grid-template-columns:repeat(2,1fr)}.grid2,.grid3,.form-grid{grid-template-columns:1fr}main{padding:18px}}@media(max-width:520px){.metrics{grid-template-columns:1fr}.toolbar,.actions{align-items:stretch;flex-direction:column}.toolbar input,.toolbar button,.actions a{width:100%}}
.section-title{display:flex;align-items:center;justify-content:space-between;margin:26px 0 12px}.section-title h2{margin:0;font-size:19px}.section-title p{margin:3px 0 0;color:var(--muted);font-size:13px}.metrics.compact{grid-template-columns:repeat(5,minmax(0,1fr))}.metric.income{border-top:4px solid #12b76a}.metric.expense{border-top:4px solid #f79009}.metric.info{border-top:4px solid #2e90fa}.metric.warning{border-top:4px solid #f04438}.metric small{display:block;margin-top:7px;color:var(--muted)}.formula{display:flex;gap:8px;align-items:center;flex-wrap:wrap;padding:14px 18px;background:#f8fafc;border:1px dashed var(--border);border-radius:14px;margin-bottom:18px}.formula strong{font-size:14px}.formula .op{color:var(--muted);font-weight:800}.money-positive{color:#067647;font-weight:800}.money-negative{color:#b42318;font-weight:800}.money-expense{color:#b54708;font-weight:700}.scroll-table{max-height:650px;overflow:auto}.scroll-table thead th{position:sticky;top:0;background:#fff;z-index:2}.table-summary{background:#f8fafc;font-weight:800}.category-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}.category-card{padding:18px;border:1px solid var(--border);border-radius:15px;background:#fff}.category-card h3{margin:0 0 12px}.category-row{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid var(--border)}.category-row:last-child{border:0}.panel-head{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;margin-bottom:12px}.panel-head h2{margin:0}.panel-head span{color:var(--muted);font-size:12px}.count-badge{background:#eef4ff;color:#3538cd;border-radius:999px;padding:5px 9px;font-weight:800;font-size:12px}@media(max-width:1100px){.metrics.compact{grid-template-columns:repeat(2,1fr)}.category-grid{grid-template-columns:1fr}}
/* Tema noturno cinza */
:root{--bg:#171a1f;--card:#23272f;--text:#f2f4f7;--muted:#a8b0bd;--border:#383e49;--dark:#0d1015;--shadow:0 12px 34px rgba(0,0,0,.28)}
body{background:linear-gradient(180deg,#15181d,#1d2128 260px);color:var(--text)}
main{background:transparent}.metric,.panel,.category-card{background:#23272f;border-color:#383e49}.quick a,.formula{background:#20242b;border-color:#383e49;color:#eef2f6}input,select,textarea{background:#1b1f25;color:#f4f6f8;border-color:#454b57}th,td{border-color:#383e49}tbody tr:hover{background:#292e37}.scroll-table thead th{background:#23272f}tr.problem{background:#3a2424}.table-summary{background:#292e37}.brand-img{object-fit:cover;padding:0;background:#111827;border:1px solid #46505f}.logo-img{object-fit:contain;padding:4px;background:#111827;border:1px solid rgba(255,255,255,.18)}.login-card,.card{background:#23272f!important;color:#f4f6f8!important}.card p,.foot{color:#a8b0bd!important}
</style></head><body><div class="layout"><aside><div class="brand"><img class="brand-logo brand-img" src="{{ url_for('company_logo') }}" alt="IRMÃOS PLATAFORMA"><div><b>IRMÃOS PLATAFORMA</b><small>Painel de Gestão · v{{ app_version }}</small></div></div><nav><div class="nav-section">Visão geral</div><a href="{{url_for('dashboard')}}">Dashboard</a><a href="{{url_for('manager_report')}}">Área do Gestor</a><div class="nav-section">Operação semanal</div><a href="{{url_for('imports_page')}}">Importações</a><a href="{{url_for('closings_page')}}">Fechamentos</a><a href="{{url_for('missing_iban_page')}}">Sem IBAN</a><a href="{{url_for('partners_page')}}">Parceiros</a><div class="nav-section">Cadastros</div><a href="{{url_for('drivers')}}">Motoristas</a><a href="{{url_for('database_page')}}">Banco de dados</a><div class="nav-section">Conta</div><a href="{{url_for('logout')}}">Sair</a></nav></aside><main><header><div><h1>{% block title %}{% endblock %}</h1><p>{% block subtitle %}{% endblock %}</p></div></header>{% for cat,msg in get_flashed_messages(with_categories=true) %}<div class="flash {{cat}}">{{msg}}</div>{% endfor %}{% block content %}{% endblock %}</main></div></body></html>"""
app.jinja_loader.mapping["login.html"] = r"""<!doctype html><html lang="pt"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Entrar · IRMÃOS PLATAFORMA</title><style>:root{--accent:#2563eb;--accent2:#06b6d4;--dark:#09111f;--muted:#667085}*{box-sizing:border-box}body{margin:0;font-family:Inter,system-ui,-apple-system,"Segoe UI",sans-serif;min-height:100vh;background:radial-gradient(circle at 20% 20%,rgba(37,99,235,.28),transparent 36%),radial-gradient(circle at 80% 80%,rgba(6,182,212,.2),transparent 34%),linear-gradient(135deg,#07101d,#111d33);display:grid;place-items:center;padding:24px}.shell{width:min(980px,100%);display:grid;grid-template-columns:1.2fr .8fr;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.12);border-radius:28px;overflow:hidden;box-shadow:0 35px 80px rgba(0,0,0,.35);backdrop-filter:blur(14px)}.hero{padding:58px;color:#fff;display:flex;flex-direction:column;justify-content:space-between;min-height:570px}.brand{display:flex;align-items:center;gap:14px}.logo{width:86px;height:86px;border-radius:20px;display:grid;place-items:center;background:#111827;font-weight:900;font-size:21px;box-shadow:0 15px 30px rgba(37,99,235,.35)}h1{font-size:46px;line-height:1.05;margin:0 0 15px;letter-spacing:-.04em}.hero p{color:#cbd5e1;font-size:17px;max-width:480px;line-height:1.6}.features{display:grid;gap:10px;color:#dbeafe;font-size:14px}.card{background:#fff;padding:46px 38px;display:flex;flex-direction:column;justify-content:center}.card h2{font-size:28px;margin:0 0 7px}.card>p{color:var(--muted);margin:0 0 28px}label{display:grid;gap:7px;margin-bottom:16px;font-size:13px;font-weight:700;color:#344054}input{width:100%;padding:13px 14px;border:1px solid #d0d5dd;border-radius:12px;font:inherit}input:focus{outline:3px solid rgba(37,99,235,.12);border-color:var(--accent)}.pass{position:relative}.pass button{position:absolute;right:7px;top:7px;border:0;background:#eef2ff;color:#344054;padding:7px 10px;border-radius:8px;box-shadow:none}.submit{width:100%;border:0;border-radius:12px;background:linear-gradient(135deg,var(--accent),#1d4ed8);color:#fff;padding:13px;font-weight:800;font-size:15px;cursor:pointer;box-shadow:0 10px 22px rgba(37,99,235,.28)}.flash{padding:11px 13px;border-radius:10px;margin-bottom:16px;background:#fee4e2;color:#b42318}.foot{text-align:center;color:#98a2b3;font-size:12px;margin-top:18px}@media(max-width:800px){.shell{grid-template-columns:1fr}.hero{min-height:auto;padding:32px}.hero h1{font-size:34px}.features{display:none}.card{padding:34px 26px}}</style></head><body><div class="shell"><section class="hero"><div class="brand"><img class="logo logo-img" src="{{ url_for('company_logo') }}" alt="IRMÃOS PLATAFORMA"><div><b style="font-size:20px">IRMÃOS PLATAFORMA</b><div style="color:#94a3b8;font-size:13px">Gestão inteligente de frotas</div></div></div><div><h1>Controle financeiro em um só lugar.</h1><p>Importe relatórios, processe pagamentos, acompanhe parceiros e prepare o fechamento semanal com segurança.</p></div><div class="features"><span>✓ TVDE e Delivery separados</span><span>✓ Combustível PRIO e consolidação por IBAN</span><span>✓ Relatórios completos para o gestor</span></div></section><form class="card" method="post"><h2>Bem-vindo</h2><p>Entre para acessar o painel administrativo.</p>{% for cat,msg in get_flashed_messages(with_categories=true) %}<div class="flash">{{msg}}</div>{% endfor %}<label>Utilizador<input name="username" autocomplete="username" required autofocus></label><label>Senha<div class="pass"><input id="password" name="password" type="password" autocomplete="current-password" required><button type="button" onclick="const p=document.getElementById('password');p.type=p.type==='password'?'text':'password';this.textContent=p.type==='password'?'Ver':'Ocultar'">Ver</button></div></label><button class="submit">Entrar no painel</button><div class="foot">Acesso exclusivo para utilizadores autorizados</div></form></div></body></html>"""
app.jinja_loader.mapping["dashboard.html"] = r"""{% extends 'base.html' %}{% block title %}Dashboard{% endblock %}{% block subtitle %}Visão clara do último fechamento processado.{% endblock %}{% block content %}<div class="quick"><a href="{{url_for('imports_page')}}">Importar relatórios</a><a href="{{url_for('closings_page')}}">Processar fechamento</a><a href="{{url_for('missing_iban_page')}}">Resolver Sem IBAN</a><a href="{{url_for('manager_report')}}">Abrir Área do Gestor</a></div><div class="section-title"><div><h2>Resumo financeiro</h2><p>Receitas, pagamentos e comissões da semana.</p></div></div><div class="metrics compact"><div class="metric income"><span>Bruto total</span><b>€ {{'%.2f'|format(stats.gross)}}</b></div><div class="metric income success"><span>Líquido a pagar</span><b>€ {{'%.2f'|format(stats.net)}}</b></div><div class="metric info"><span>Comissão da empresa</span><b>€ {{'%.2f'|format(stats.commission)}}</b></div><div class="metric info"><span>Transferências</span><b>{{stats.payments}}</b></div><div class="metric warning"><span>Sem IBAN</span><b>{{stats.missing_iban}}</b></div></div><div class="section-title"><div><h2>Descontos e ajustes</h2><p>Valores que reduzem ou ajustam o pagamento.</p></div></div><div class="metrics compact"><div class="metric expense"><span>Combustível</span><b>€ {{'%.2f'|format(stats.fuel)}}</b>{% for f in fuel_breakdown %}<small>{{f.source_file|replace('.xlsx','')|replace('.xls','')}}: € {{'%.2f'|format(f.total)}}</small>{% endfor %}</div><div class="metric expense"><span>Descontos gerais</span><b>€ {{'%.2f'|format(stats.discount)}}</b></div><div class="metric income"><span>Reembolsos</span><b>€ {{'%.2f'|format(stats.reimbursement)}}</b></div></div><div class="grid2"><section class="panel"><div class="panel-head"><div><h2>Bruto por origem</h2><span>Comparação entre plataformas.</span></div></div><canvas id="originChart"></canvas></section><section class="panel"><div class="panel-head"><div><h2>Composição dos ajustes</h2><span>Comissão, combustível, descontos e reembolsos.</span></div></div><canvas id="moneyChart"></canvas></section></div><section class="panel"><div class="panel-head"><div><h2>Fechamentos recentes</h2><span>Histórico das últimas semanas processadas.</span></div></div><div class="table-wrap"><table><thead><tr><th>ID</th><th>Semana</th><th>Data</th><th>Status</th><th></th></tr></thead><tbody>{% for r in recent %}<tr><td>#{{r.id}}</td><td>{{r.label}}</td><td>{{r.created_at}}</td><td>{{r.status}}</td><td><a class="link" href="{{url_for('closing_detail',closing_id=r.id)}}">Abrir</a></td></tr>{% else %}<tr><td colspan="5">Nenhum fechamento processado.</td></tr>{% endfor %}</tbody></table></div></section><script>new Chart(document.getElementById('originChart'),{type:'bar',data:{labels:{{origin_labels|tojson}},datasets:[{label:'Bruto (€)',data:{{origin_values|tojson}}}]},options:{responsive:true,plugins:{legend:{display:false}}}});new Chart(document.getElementById('moneyChart'),{type:'doughnut',data:{labels:['Comissão','Combustível','Descontos','Reembolsos'],datasets:[{data:[{{stats.commission}},{{stats.fuel}},{{stats.discount}},{{stats.reimbursement}}]}]},options:{responsive:true}});</script>{% endblock %}"""
app.jinja_loader.mapping["manager.html"] = r"""{% extends 'base.html' %}{% block title %}Área do Gestor{% endblock %}{% block subtitle %}Relatório administrativo do último fechamento: {{closing.label if closing else 'nenhum fechamento'}}.{% endblock %}{% block content %}{% if closing %}<div class="actions"><a class="button" href="{{url_for('manager_excel',closing_id=closing.id)}}">Baixar relatório completo</a></div><div class="section-title"><div><h2>Resultado da semana</h2><p>Principais valores financeiros do fechamento.</p></div></div><div class="metrics compact"><div class="metric income"><span>Bruto total</span><b>€ {{'%.2f'|format(stats.gross)}}</b></div><div class="metric income success"><span>Líquido a pagar</span><b>€ {{'%.2f'|format(stats.net)}}</b></div><div class="metric info"><span>Comissão da empresa</span><b>€ {{'%.2f'|format(stats.commission)}}</b></div><div class="metric info"><span>Comissão dos parceiros</span><b>€ {{'%.2f'|format(stats.partner_commission)}}</b></div></div><div class="formula"><strong>Bruto</strong><span class="op">−</span><strong>Combustível</strong><span class="op">−</span><strong>Descontos</strong><span class="op">−</span><strong>Dinheiro em mãos</strong><span class="op">+</span><strong>Reembolsos</strong><span class="op">=</span><strong class="money-positive">Líquido a pagar</strong></div><div class="section-title"><div><h2>Descontos e ajustes</h2><p>Itens que alteram o valor final dos motoristas.</p></div></div><div class="metrics compact"><div class="metric expense"><span>Combustível</span><b>€ {{'%.2f'|format(stats.fuel)}}</b>{% for f in fuel_breakdown %}<small>{{f.source_file|replace('.xlsx','')|replace('.xls','')}}: € {{'%.2f'|format(f.total)}}</small>{% endfor %}</div><div class="metric expense"><span>Descontos gerais</span><b>€ {{'%.2f'|format(stats.discount)}}</b></div><div class="metric expense"><span>Dinheiro em mãos</span><b>€ {{'%.2f'|format(stats.cash)}}</b></div><div class="metric income"><span>Reembolsos</span><b>€ {{'%.2f'|format(stats.reimbursement)}}</b></div></div><div class="section-title"><div><h2>Operação da semana</h2><p>Alertas e divisão das atividades.</p></div></div><div class="metrics compact"><div class="metric warning"><span>Pessoas negativas</span><b>{{stats.negatives}}</b></div><div class="metric warning"><span>Sem IBAN</span><b>{{stats.missing_iban}}</b></div></div><div class="category-grid"><div class="category-card"><h3>TVDE</h3><div class="category-row"><span>Bruto</span><b>€ {{'%.2f'|format(stats.tvde)}}</b></div></div><div class="category-card"><h3>Delivery</h3><div class="category-row"><span>Bruto</span><b>€ {{'%.2f'|format(stats.delivery)}}</b></div></div></div><div class="section-title"><div><h2>Análises detalhadas</h2><p>Parceiros e motoristas que precisam de atenção.</p></div></div><div class="grid2"><section class="panel table-wrap"><div class="panel-head"><div><h2>Ganhos por parceiro</h2><span>Ordenado pelo maior valor bruto.</span></div><span class="count-badge">{{partners|length}} parceiros</span></div><table><thead><tr><th>Parceiro</th><th>Motoristas</th><th>Bruto</th><th>Comissão</th><th>Líquido</th></tr></thead><tbody>{% for r in partners %}<tr><td><b>{{r.partner}}</b></td><td>{{r.drivers}}</td><td>€ {{'%.2f'|format(r.gross)}}</td><td class="money-expense">€ {{'%.2f'|format(r.partner_commission)}}</td><td class="money-positive">€ {{'%.2f'|format(r.net)}}</td></tr>{% else %}<tr><td colspan="5">Nenhum parceiro.</td></tr>{% endfor %}</tbody></table></section><section class="panel table-wrap scroll-table"><div class="panel-head"><div><h2>Pessoas negativas</h2><span>Saldo final abaixo de zero.</span></div><span class="count-badge">{{negatives|length}}</span></div><table><thead><tr><th>Motorista</th><th>Origem</th><th>Parceiro</th><th>Saldo final</th></tr></thead><tbody>{% for r in negatives %}<tr class="problem"><td><b>{{r.driver_name}}</b></td><td>{{r.origins}}</td><td>{{r.partner or '—'}}</td><td class="money-negative">€ {{'%.2f'|format(r.net)}}</td></tr>{% else %}<tr><td colspan="4">Nenhuma pessoa negativa.</td></tr>{% endfor %}</tbody></table></section></div><section class="panel table-wrap"><div class="panel-head"><div><h2>Sem IBAN</h2><span>Motoristas que impedem a inclusão no XML.</span></div><span class="count-badge">{{missing|length}}</span></div><table><thead><tr><th>Motorista</th><th>Origem</th><th>Parceiro</th><th>Líquido</th></tr></thead><tbody>{% for r in missing %}<tr class="problem"><td><b>{{r.driver_name}}</b></td><td>{{r.origins}}</td><td>{{r.partner or '—'}}</td><td>€ {{'%.2f'|format(r.net)}}</td></tr>{% else %}<tr><td colspan="4">Nenhum motorista sem IBAN.</td></tr>{% endfor %}</tbody></table></section>{% else %}<section class="panel"><h2>Nenhum fechamento disponível</h2><p>Importe os relatórios e processe uma semana para visualizar o relatório do gestor.</p></section>{% endif %}{% endblock %}"""

app.secret_key = os.getenv("SECRET_KEY", "troque-esta-chave-em-producao")
app.config["MAX_CONTENT_LENGTH"] = 80 * 1024 * 1024
COMPANY_NAME = os.getenv("COMPANY_NAME", "IRMÃOS PLATAFORMA")
APP_VERSION = "1.2.6"
ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")


def db() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH, timeout=30)
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=30000")
    con.row_factory = sqlite3.Row
    return con


def norm(v: Any) -> str:
    s = "" if v is None else str(v).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).upper()


def is_uber_tvde_origin(v: Any) -> bool:
    """Aceita as duas formas usadas no projeto: UBER_TVDE e UBER TVDE."""
    key = norm(v).replace("_", " ")
    return "UBER TVDE" in key


def clean_identifier(v: Any) -> str:
    return re.sub(r"[^A-Z0-9@.+_-]", "", norm(v))


def clean_card_number(v: Any) -> str:
    """Normaliza cartões PRIO e do banco para o mesmo formato numérico."""
    if v is None:
        return ""
    s = str(v).strip().replace("'", "")
    # Excel pode devolver cartões numéricos como 7824... .0 ou notação científica.
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    elif re.fullmatch(r"[+-]?\d+(?:\.\d+)?[Ee][+-]?\d+", s):
        try:
            s = format(float(s), ".0f")
        except ValueError:
            pass
    return re.sub(r"\D", "", s)


def money(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").replace(" ", "")
    s = s.replace("−", "-").replace("–", "-")
    negative_parentheses = s.startswith("(") and s.endswith(")")
    if negative_parentheses:
        s = s[1:-1]
    if not s or s.upper() in {"-", "NAN", "NONE", "NULL"}:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        value = float(s)
        return -value if negative_parentheses else value
    except ValueError:
        return 0.0



@app.template_filter("format")
def locale_format(fmt: str, value: Any) -> str:
    """Mantém os templates existentes, mas exibe números no padrão pt-PT."""
    try:
        number = float(value or 0)
        decimals = 2 if ".2" in str(fmt) else 0
        rendered = f"{number:,.{decimals}f}"
        return rendered.replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        try:
            return str(fmt) % value
        except Exception:
            return str(value)


def style_workbook(book) -> None:
    """Padroniza todos os relatórios Excel sem alterar os dados."""
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D9E2F0")
    money_words = {"bruto","dinheiro","comissao","combustivel","desconto","reembolso","imediata","taxa","liquido","total"}
    for ws in book.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)
        for col in ws.columns:
            letter = col[0].column_letter
            width = min(max(max(len(str(c.value or "")) for c in col) + 2, 12), 34)
            ws.column_dimensions[letter].width = width
        for idx, cell in enumerate(ws[1], 1):
            name = str(cell.value or "").lower()
            if any(word in name for word in money_words):
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, idx).number_format = '€ #,##0.00;[Red]-€ #,##0.00'
        if ws.max_row > 1:
            for cell in ws[ws.max_row]:
                if str(cell.value or "").upper() in {"TOTAL", "TOTAIS"}:
                    for c in ws[ws.max_row]:
                        c.fill = total_fill
                        c.font = Font(bold=True)
                    break

def init_db() -> None:
    # O Render pode iniciar mais de um processo ao mesmo tempo.
    # As tentativas abaixo evitam falha temporária por bloqueio do SQLite.
    last_error = None
    for attempt in range(5):
        try:
            with db() as con:
                con.executescript("""
        CREATE TABLE IF NOT EXISTS drivers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            external_id TEXT,
            name TEXT NOT NULL,
            city TEXT,
            company TEXT,
            iban TEXT,
            bank_color TEXT,
            commission_owner REAL DEFAULT 0,
            partner TEXT,
            partner_commission REAL DEFAULT 0,
            percentage REAL DEFAULT 0,
            fuel_card TEXT,
            rental_label TEXT,
            discount REAL DEFAULT 0,
            reimbursement REAL DEFAULT 0,
            immediate REAL DEFAULT 0,
            observation TEXT,
            updated_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_driver_external ON drivers(external_id);
        CREATE INDEX IF NOT EXISTS idx_driver_iban ON drivers(iban);
        CREATE TABLE IF NOT EXISTS mappings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT UNIQUE,
            origin_ref TEXT,
            identifier_column TEXT,
            first_name_column TEXT,
            last_name_column TEXT,
            value_column TEXT,
            cash_column TEXT,
            reimbursement_1 TEXT,
            reimbursement_2 TEXT,
            reimbursement_3 TEXT,
            reimbursement_4 TEXT
        );
        CREATE TABLE IF NOT EXISTS imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            sha256 TEXT UNIQUE,
            kind TEXT,
            status TEXT,
            rows_count INTEGER DEFAULT 0,
            message TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS raw_earnings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER,
            filename TEXT,
            origin_ref TEXT,
            identifier TEXT,
            display_name TEXT,
            gross REAL DEFAULT 0,
            cash REAL DEFAULT 0,
            reimbursement REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS fuel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER,
            card_number TEXT,
            amount REAL DEFAULT 0,
            source_file TEXT
        );
        CREATE TABLE IF NOT EXISTS closings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT,
            created_at TEXT,
            status TEXT
        );
        CREATE TABLE IF NOT EXISTS closing_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            closing_id INTEGER,
            driver_id INTEGER,
            driver_name TEXT,
            iban TEXT,
            bank_color TEXT,
            origins TEXT,
            gross REAL,
            cash REAL,
            commission REAL,
            fuel REAL,
            discount REAL,
            reimbursement REAL,
            immediate REAL,
            bank_fee REAL,
            net_before_group REAL,
            group_total REAL
        );
        """)
            break
        except sqlite3.OperationalError as exc:
            last_error = exc
            time.sleep(1 + attempt)
    else:
        raise RuntimeError(f"Não foi possível inicializar o banco: {last_error}")

    if SEED_XLSX.exists():
        with db() as con:
            count = con.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
        if count == 0:
            try:
                import_master_workbook(SEED_XLSX)
            except sqlite3.IntegrityError:
                # Outro processo pode ter terminado a importação primeiro.
                pass


def import_master_workbook(path: Path) -> tuple[int, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Banco_de_Dados"]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    drivers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[idx.get("MOTORISTA_BD", 2)] if len(row) > 2 else None
        if not name:
            continue
        drivers.append((
            str(row[idx.get("ID", 1)] or ""), str(name), str(row[idx.get("Cidade", 4)] or ""),
            str(row[idx.get("Companhia", 5)] or ""), str(row[idx.get("IBAN", 6)] or ""),
            str(row[idx.get("Banco", 8)] or ""), money(row[idx.get("Comissao_Muryllo", 9)]),
            str(row[idx.get("Parceiro_2", 10)] or ""), money(row[idx.get("Comissao_2", 11)]),
            money(row[idx.get("Porcentagem", 12)]), str(row[idx.get("Numero_Cartao_Combustivel", 13)] or ""),
            str(row[idx.get("Aluguel", 15)] or ""), money(row[idx.get("Desconto_(-)", 16)]),
            money(row[idx.get("Reembolso_(+)", 17)]), money(row[idx.get("Imediata_(-)", 18)]),
            str(row[idx.get("Observação", 19)] or ""), datetime.now().isoformat(timespec="seconds")
        ))
    with db() as con:
        con.execute("DELETE FROM drivers")
        con.executemany("""INSERT INTO drivers(external_id,name,city,company,iban,bank_color,commission_owner,partner,partner_commission,percentage,fuel_card,rental_label,discount,reimbursement,immediate,observation,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", drivers)

    ws2 = wb["Arquivos_Pgto"]
    mappings = []
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        vals = list(r[:11]) + [None] * (11-len(r[:11]))
        mappings.append(tuple("" if v is None else str(v) for v in vals[:11]))
    with db() as con:
        con.execute("DELETE FROM mappings")
        con.executemany("""INSERT INTO mappings(file_name,origin_ref,identifier_column,first_name_column,last_name_column,value_column,cash_column,reimbursement_1,reimbursement_2,reimbursement_3,reimbursement_4)
        VALUES(?,?,?,?,?,?,?,?,?,?,?)""", mappings)
    return len(drivers), len(mappings)


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def read_csv_flexible(path: Path) -> pd.DataFrame:
    last_error = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        for sep in (None, ",", ";", "\t"):
            try:
                kwargs = {"encoding": enc, "dtype": str, "keep_default_na": False}
                if sep is None:
                    kwargs.update({"sep": None, "engine": "python"})
                else:
                    kwargs.update({"sep": sep})
                df = pd.read_csv(path, **kwargs)
                if len(df.columns) >= 2:
                    df.columns = [str(c).strip() for c in df.columns]
                    return df
            except Exception as exc:
                last_error = exc
    raise ValueError(f"Não foi possível ler o CSV: {last_error}")


def find_mapping(filename: str, columns: list[str]):
    stem = norm(Path(filename).stem)
    with db() as con:
        maps = con.execute("SELECT * FROM mappings ORDER BY LENGTH(file_name) DESC").fetchall()
    for m in maps:
        if norm(m["file_name"]) == stem or norm(m["file_name"]) in stem:
            return m
    # fallback only when exact file mapping is absent
    nc = {norm(c): c for c in columns}
    if "EMAIL" in nc and any("ADJUSTED EARNINGS" in k for k in nc):
        return {"origin_ref":"BOLT_FOOD","identifier_column":nc["EMAIL"],"first_name_column":nc.get("FIRST NAME",""),"last_name_column":nc.get("LAST NAME",""),"value_column":next(v for k,v in nc.items() if "ADJUSTED EARNINGS" in k and "VAT" in k),"cash_column":"-","reimbursement_1":"-","reimbursement_2":"-","reimbursement_3":"-","reimbursement_4":"-"}
    if "TELEMOVEL" in nc or "TELEMÓVEL" in columns:
        key = nc.get("TELEMOVEL", "Telemóvel")
        val = next((v for k,v in nc.items() if "PAGAMENTO PREVISTO" in k), "")
        return {"origin_ref":"BOLT_TVDE","identifier_column":key,"first_name_column":nc.get("MOTORISTA",""),"last_name_column":"-","value_column":val,"cash_column":"-","reimbursement_1":"-","reimbursement_2":"-","reimbursement_3":"-","reimbursement_4":"-"}
    if "UUID DO MOTORISTA" in nc:
        return {"origin_ref":"UBER EATS","identifier_column":nc["UUID DO MOTORISTA"],"first_name_column":nc.get("NOME PROPRIO DO MOTORISTA",""),"last_name_column":nc.get("APELIDO DO MOTORISTA",""),"value_column":nc.get("PAGO A SI",""),"cash_column":next((v for k,v in nc.items() if "DINHEIRO RECEBIDO" in k),"-"),"reimbursement_1":"-","reimbursement_2":"-","reimbursement_3":"-","reimbursement_4":"-"}
    return None


def canonical_column(v: Any) -> str:
    """Compara cabeçalhos ignorando acentos, espaços e pontuação.

    Os CSVs da Uber variam entre, por exemplo, ``Pago a si:Saldo`` e
    ``Pago a si : Saldo``. A comparação anterior considerava esses nomes
    diferentes e os campos de reembolso acabavam lidos como zero.
    """
    return re.sub(r"[^A-Z0-9]", "", norm(v))


def _canonical_report_header(v: Any) -> str:
    # O pandas acrescenta .1, .2... quando existem cabeçalhos repetidos.
    # Esse sufixo não faz parte do nome configurado em Arquivos_Pgto.
    text = re.sub(r"\.\d+$", "", str(v or "").strip())
    return canonical_column(text)


def resolve_configured_column(columns: list[Any], configured: str) -> Any | None:
    """Resolve somente o cabeçalho configurado em Arquivos_Pgto.

    Não tenta adivinhar colunas por palavras. Apenas ignora diferenças de
    acentos, espaços, pontuação, quebras de linha e o sufixo .1 do pandas.
    """
    if not configured or str(configured).strip() == "-":
        return None
    if configured in columns:
        return configured
    target = _canonical_report_header(configured)
    if not target:
        return None
    for column in columns:
        if _canonical_report_header(column) == target:
            return column
    return None


def col_value(row: pd.Series, col: str) -> Any:
    resolved = resolve_configured_column(list(row.index), col)
    return row[resolved] if resolved is not None else 0


def correct_origin_by_file_and_columns(filename: str, mapping, columns: list[str]):
    """Corrige a origem usando o nome real do relatório e seu cabeçalho.

    Arquivos de empresas/cidades podem ter conteúdo idêntico e ainda assim são
    relatórios distintos. Também diferencia Uber TVDE de Uber Eats, que usam
    cabeçalhos semelhantes.
    """
    name = norm(Path(filename).stem)
    cols = {norm(c) for c in columns}
    result = dict(mapping)

    configured_origin = norm(result.get("origin_ref", ""))

    if "RELATORIO BOLT FOOD" in name:
        result["origin_ref"] = "BOLT_FOOD"
    elif "RELATORIO BOLT" in name:
        result["origin_ref"] = "BOLT_TVDE"
    elif "TVDE" in name and "BOLT" not in name:
        # Aceita qualquer variação do nome do relatório Uber TVDE, inclusive
        # nomes com data, cidade ou empresa antes/depois de TVDE.
        result["origin_ref"] = "UBER_TVDE"
    elif "TVDE" in configured_origin:
        # O mapeamento da folha Arquivos_Pgto é a fonte principal. Não deixa
        # o cabeçalho UUID, comum ao Uber TVDE e Uber Eats, trocar TVDE por Eats.
        result["origin_ref"] = result.get("origin_ref") or "UBER_TVDE"
    elif "UUID DO MOTORISTA" in cols:
        # Apenas quando o mapeamento não diz TVDE, os relatórios com UUID são
        # tratados como Uber Eats/Delivery.
        result["origin_ref"] = "UBER EATS"

    return result


def ensure_provisional_driver(identifier: Any, display_name: Any, origin: Any) -> None:
    """Cria cadastro mínimo para motorista novo aparecer em Sem IBAN.

    O cadastro completo continua podendo ser importado depois pelo Banco_de_Dados.
    Não substitui nem altera motorista já existente.
    """
    external_id = str(identifier or "").strip()
    clean_ext = clean_identifier(external_id)
    if not clean_ext:
        return
    name = str(display_name or "").strip() or external_id
    with db() as con:
        existing = con.execute("SELECT id,external_id FROM drivers").fetchall()
        if any(clean_identifier(r["external_id"]) == clean_ext for r in existing):
            return
        con.execute(
            """INSERT INTO drivers(
                external_id,name,city,company,iban,bank_color,commission_owner,
                partner,partner_commission,percentage,fuel_card,rental_label,
                discount,reimbursement,immediate,observation,updated_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (external_id, name, "", str(origin or ""), "", "YELLOW", 0,
             "", 0, 0, "", "", 0, 0, 0,
             "Cadastro automático: motorista novo encontrado no relatório; complete os dados.", now()),
        )


def process_report_file(path: Path, import_id: int) -> tuple[int, str]:
    df = read_csv_flexible(path)
    mapping = find_mapping(path.name, list(df.columns))
    if not mapping:
        raise ValueError("O nome do ficheiro não está configurado em Arquivos_Pgto e o cabeçalho não foi reconhecido.")
    mapping = correct_origin_by_file_and_columns(path.name, mapping, list(df.columns))
    origin = str(mapping["origin_ref"])
    count = 0
    rows = []
    for _, r in df.iterrows():
        identifier = str(col_value(r, mapping["identifier_column"])).strip()
        if not identifier:
            continue
        first = str(col_value(r, mapping["first_name_column"])).strip() if mapping["first_name_column"] != "-" else ""
        last = str(col_value(r, mapping["last_name_column"])).strip() if mapping["last_name_column"] != "-" else ""
        gross = money(col_value(r, mapping["value_column"]))
        cash = money(col_value(r, mapping["cash_column"]))

        # Uber TVDE: o campo "Pago a si" inclui ajustes/reembolsos e, por isso,
        # não deve ser usado como bruto. O bruto correto é "Pago a si : Os seus
        # rendimentos"; o reembolso entra separadamente mais abaixo.
        if is_uber_tvde_origin(origin):
            tvde_gross_column = resolve_configured_column(
                list(r.index), "Pago a si : Os seus rendimentos"
            )
            if tvde_gross_column is not None:
                gross = money(r[tvde_gross_column])
            # Relatórios TVDE não entram em Dinheiro em mãos.
            cash = 0.0

        # Nos relatórios Uber Eats, o dinheiro recebido costuma vir negativo.
        # Usamos o valor absoluto: entra na base da comissão e é abatido no final.
        if "EATS" in norm(origin):
            cash = abs(cash)
        else:
            cash = 0.0
        # No Uber TVDE, as colunas de reembolso são definidas explicitamente
        # na aba Arquivos_Pgto (Reembolso_1 a Reembolso_4). O sistema não
        # procura palavras parecidas e não tenta adivinhar cabeçalhos.
        reimb = 0.0
        if is_uber_tvde_origin(origin):
            used_reimbursement_columns = set()
            for key in ("reimbursement_1", "reimbursement_2", "reimbursement_3", "reimbursement_4"):
                configured = str(mapping.get(key, "") or "").strip()
                canonical = _canonical_report_header(configured)
                if not configured or configured == "-" or not canonical or canonical in used_reimbursement_columns:
                    continue
                used_reimbursement_columns.add(canonical)
                resolved = resolve_configured_column(list(r.index), configured)
                if resolved is not None:
                    reimb += money(r[resolved])

            # Correção para o formato real exportado pela Uber. Alguns bancos
            # antigos têm a coluna de reembolso vazia ou com nome desatualizado.
            # Nesse caso usa exatamente a coluna real de portagem, sem somá-la
            # duas vezes caso ela já tenha sido encontrada pelo mapeamento.
            exact_reimbursement_name = "Pago a si:Saldo da viagem:Reembolsos:Portagem"
            exact_canonical = _canonical_report_header(exact_reimbursement_name)
            if exact_canonical not in used_reimbursement_columns:
                resolved = resolve_configured_column(list(r.index), exact_reimbursement_name)
                if resolved is not None:
                    reimb += money(r[resolved])
        rows.append((import_id, path.name, origin, identifier, (first + " " + last).strip(), gross, cash, reimb))
        count += 1
    if not rows and len(df.index) == 0:
        return 0, "Ficheiro válido, mas sem linhas de pagamento."
    with db() as con:
        con.executemany("INSERT INTO raw_earnings(import_id,filename,origin_ref,identifier,display_name,gross,cash,reimbursement) VALUES(?,?,?,?,?,?,?,?)", rows)
    # Motoristas novos passam a aparecer imediatamente em Sem IBAN, mesmo
    # antes de existirem no ficheiro Banco_de_Dados.
    for raw in rows:
        ensure_provisional_driver(raw[3], raw[4], raw[2])
    reimbursement_total = sum(money(r[7]) for r in rows)
    if "TVDE" in norm(origin):
        return count, f"{origin}: {count} linhas · reembolsos € {reimbursement_total:.2f}"
    return count, f"{origin}: {count} linhas"


def process_fuel_file(path: Path, import_id: int) -> tuple[int, str]:
    """Lê relatórios PRIO cujo cabeçalho real começa normalmente na 4.ª linha."""
    workbook = pd.ExcelFile(path)
    grouped = defaultdict(float)
    sheets_read = 0

    for sheet in workbook.sheet_names:
        # Localiza automaticamente a linha do cabeçalho (o modelo original usa header=3).
        preview = pd.read_excel(path, sheet_name=sheet, header=None, nrows=12, dtype=str).fillna("")
        header_row = None
        for i, row in preview.iterrows():
            names = [norm(v) for v in row.tolist()]
            has_card = any("CARTAO" in x or x == "PAN" for x in names)
            has_total = any(x == "TOTAL" or "VALOR" in x or "MONTANTE" in x or "DEBITO" in x for x in names)
            if has_card and has_total:
                header_row = int(i)
                break
        if header_row is None:
            continue

        df = pd.read_excel(path, sheet_name=sheet, header=header_row, dtype=str).fillna("")
        cols = list(df.columns)
        card_col = next((c for c in cols if "CARTAO" in norm(c) or norm(c) == "PAN"), None)
        amount_col = next((c for c in cols if norm(c) == "TOTAL"), None)
        if amount_col is None:
            amount_col = next((c for c in cols if any(x in norm(c) for x in ("VALOR", "MONTANTE", "DEBITO"))), None)
        if card_col is None or amount_col is None:
            continue

        sheets_read += 1
        for _, r in df.iterrows():
            card = clean_card_number(r.get(card_col, ""))
            amount = money(r.get(amount_col, ""))
            if card and amount != 0:
                grouped[card] += abs(amount)

    if not grouped:
        raise ValueError("Não foi possível encontrar despesas PRIO. O arquivo precisa conter as colunas CARTÃO e TOTAL, normalmente a partir da 4.ª linha.")

    with db() as con:
        con.executemany(
            "INSERT INTO fuel(import_id,card_number,amount,source_file) VALUES(?,?,?,?)",
            [(import_id, card, round(amount, 2), path.name) for card, amount in grouped.items()],
        )
    total = sum(grouped.values())
    return len(grouped), f"{len(grouped)} cartões PRIO consolidados em {sheets_read} aba(s), total € {total:.2f}"


def match_driver(identifier: str, origin: str, display_name: str):
    ident = clean_identifier(identifier)
    with db() as con:
        drivers = con.execute("SELECT * FROM drivers").fetchall()
    if "BOLT_FOOD" in norm(origin):
        # the current source workbook stores Bolt Food identifiers in ID where available
        for d in drivers:
            if clean_identifier(d["external_id"]) == ident:
                return d
    if "BOLT_TVDE" in norm(origin):
        digits = re.sub(r"\D", "", identifier)
        for d in drivers:
            dd = re.sub(r"\D", "", str(d["external_id"] or ""))
            if digits and dd and (digits == dd or digits[-9:] == dd[-9:]):
                return d
    for d in drivers:
        if clean_identifier(d["external_id"]) == ident:
            return d
    n = norm(display_name)
    if n:
        exact = [d for d in drivers if norm(d["name"]) == n]
        if len(exact) == 1:
            return exact[0]
    return None


def build_closing(label: str) -> int:
    """Processa o fechamento sem deixar uma linha inválida derrubar a aplicação."""
    created_at = datetime.now().isoformat(timespec="seconds")
    with db() as con:
        cur = con.execute(
            "INSERT INTO closings(label,created_at,status) VALUES(?,?,?)",
            (label, created_at, "PROCESSANDO"),
        )
        closing_id = int(cur.lastrowid)

    try:
        with db() as con:
            earnings = [dict(r) for r in con.execute("SELECT * FROM raw_earnings").fetchall()]
            fuel_rows = [dict(r) for r in con.execute(
                "SELECT card_number,SUM(amount) amount FROM fuel GROUP BY card_number"
            ).fetchall()]
            drivers = [dict(r) for r in con.execute("SELECT * FROM drivers").fetchall()]

        fuel_map = {
            clean_card_number(r.get("card_number")): money(r.get("amount"))
            for r in fuel_rows
            if clean_card_number(r.get("card_number"))
        }

        by_external = defaultdict(list)
        by_name = defaultdict(list)
        for d in drivers:
            ext = clean_identifier(d.get("external_id"))
            if ext:
                by_external[ext].append(d)
            name_key = norm(d.get("name"))
            if name_key:
                by_name[name_key].append(d)

        def find_driver(identifier: Any, origin: Any, display_name: Any):
            ident = clean_identifier(identifier)
            origin_key = norm(origin)
            if "BOLT_TVDE" in origin_key:
                digits = re.sub(r"\D", "", str(identifier or ""))
                if digits:
                    for d in drivers:
                        dd = re.sub(r"\D", "", str(d.get("external_id") or ""))
                        if dd and (digits == dd or digits[-9:] == dd[-9:]):
                            return d
            if ident and len(by_external.get(ident, [])) == 1:
                return by_external[ident][0]
            name_key = norm(display_name)
            if name_key and len(by_name.get(name_key, [])) == 1:
                return by_name[name_key][0]
            return None

        # Mantém TVDE separado de Delivery desde a leitura. Assim o mesmo
        # motorista/IBAN pode ter um pagamento TVDE e outro Delivery sem misturar.
        aggregated: dict[tuple[int, str], dict[str, Any]] = {}
        unmatched = 0
        invalid_rows = 0
        for e in earnings:
            try:
                driver = find_driver(e.get("identifier"), e.get("origin_ref"), e.get("display_name"))
                if not driver:
                    ensure_provisional_driver(e.get("identifier"), e.get("display_name"), e.get("origin_ref"))
                    ident = clean_identifier(e.get("identifier"))
                    with db() as lookup_con:
                        candidates = lookup_con.execute("SELECT * FROM drivers").fetchall()
                    driver = next((dict(r) for r in candidates if clean_identifier(r["external_id"]) == ident), None)
                if not driver:
                    unmatched += 1
                    continue
                driver_id = int(driver["id"])
                origin = str(e.get("origin_ref") or "SEM ORIGEM")
                origin_norm = norm(origin)
                category = "TVDE" if "TVDE" in origin_norm else "DELIVERY"
                key = (driver_id, category)
                a = aggregated.setdefault(key, {
                    "driver": driver,
                    "category": category,
                    "gross": 0.0,
                    "cash": 0.0,
                    "report_reimb": 0.0,
                    "origins": set(),
                })
                a["gross"] += money(e.get("gross"))
                row_cash = money(e.get("cash"))
                if "EATS" in origin_norm:
                    a["cash"] += abs(row_cash)
                a["report_reimb"] += money(e.get("reimbursement"))
                a["origins"].add(origin)
            except Exception:
                invalid_rows += 1
                app.logger.exception("Linha ignorada durante o fechamento: %r", e)

        items = []
        # Combustível, desconto e imediata são dados fixos do motorista e devem
        # ser aplicados somente uma vez por fechamento, mesmo quando ele tem
        # ganhos em TVDE e Delivery. Damos prioridade ao TVDE; se não existir,
        # aplicamos no Delivery.
        driver_categories = defaultdict(set)
        for (driver_id, category) in aggregated.keys():
            driver_categories[driver_id].add(category)
        deduction_category = {
            driver_id: ("TVDE" if "TVDE" in categories else "DELIVERY")
            for driver_id, categories in driver_categories.items()
        }

        # O mesmo motorista pode existir em duas linhas do banco (Uber e Bolt)
        # com o mesmo desconto cadastrado nas duas. Esses valores representam
        # uma única cobrança. Consolidamos pelo nome lógico e usamos o maior
        # valor cadastrado, aplicando-o somente uma vez no fechamento.
        logical_fixed_values = {}
        for drv in drivers:
            logical_key = norm(drv.get("name")) or f"ID:{drv.get('id')}"
            current = logical_fixed_values.setdefault(logical_key, {
                "discount": 0.0, "immediate": 0.0, "manual_reimbursement": 0.0
            })
            current["discount"] = max(current["discount"], money(drv.get("discount")))
            current["immediate"] = max(current["immediate"], money(drv.get("immediate")))
            current["manual_reimbursement"] = max(current["manual_reimbursement"], money(drv.get("reimbursement")))

        # Consolidação e taxa bancária são feitas por categoria + IBAN.
        iban_groups = defaultdict(float)
        temp = []
        # TVDE é processado primeiro para receber os descontos fixos quando
        # o mesmo motorista também aparece em Delivery.
        applied_fuel_cards = set()
        applied_fixed_keys = set()
        ordered_aggregated = sorted(
            aggregated.values(),
            key=lambda x: (0 if x["category"] == "TVDE" else 1, int(x["driver"]["id"]))
        )
        for a in ordered_aggregated:
            d = a["driver"]
            category = a["category"]
            origins_set = a["origins"]
            origins = ", ".join(sorted(origins_set))
            gross = money(a["gross"])
            cash = money(a["cash"])
            pct = money(d.get("percentage"))
            if pct > 1:
                pct = pct / 100.0
            pct = max(0.0, min(pct, 1.0))
            is_eats = any("EATS" in norm(x) for x in origins_set)
            is_tvde = category == "TVDE"
            commission_base = gross + cash if is_eats else gross
            commission = commission_base * pct
            apply_fixed_deductions = category == deduction_category.get(int(d["id"]), category)
            card_key = clean_card_number(d.get("fuel_card"))
            logical_driver_key = norm(d.get("name")) or f"ID:{d['id']}"
            fixed_key = (logical_driver_key, category if not apply_fixed_deductions else "FIXED")
            # O mesmo cartão pode estar repetido em duas linhas do banco (por
            # exemplo, uma linha TVDE e outra Delivery). O valor PRIO é abatido
            # uma única vez por número de cartão em todo o fechamento.
            fuel_value = 0.0
            if apply_fixed_deductions and card_key and card_key not in applied_fuel_cards:
                fuel_value = fuel_map.get(card_key, 0.0)
                applied_fuel_cards.add(card_key)
            # Desconto e imediata também são fixos do motorista. Se o cadastro
            # estiver duplicado por plataforma, aplica apenas uma vez pelo nome.
            use_fixed = apply_fixed_deductions and fixed_key not in applied_fixed_keys
            logical_values = logical_fixed_values.get(logical_driver_key, {})
            discount = money(logical_values.get("discount")) if use_fixed else 0.0
            immediate = money(logical_values.get("immediate")) if use_fixed else 0.0
            if use_fixed:
                applied_fixed_keys.add(fixed_key)
            reimbursement = (
                money(logical_values.get("manual_reimbursement")) + money(a["report_reimb"])
            ) if is_tvde else 0.0
            # Uber Eats: dinheiro em mãos entra na base da comissão e é abatido
            # integralmente no final, porque já está com o motorista.
            if is_eats:
                net = (gross + cash) - commission - fuel_value - discount - immediate - cash
            else:
                net = gross - commission - fuel_value - discount - immediate + reimbursement
            iban = re.sub(r"\s+", "", str(d.get("iban") or "")).upper()
            temp.append({
                "driver": d, "category": category, "origins": origins,
                "gross": gross, "cash": cash, "commission": commission,
                "fuel": fuel_value, "discount": discount,
                "reimbursement": reimbursement, "immediate": immediate,
                "net": net, "iban": iban,
            })
            if iban:
                iban_groups[(category, iban)] += net

        fee_groups = {
            (t["category"], t["iban"]) for t in temp
            if t["iban"] and norm(t["driver"].get("bank_color")) == "AZUL"
        }
        for group_key in fee_groups:
            iban_groups[group_key] -= 1.25

        fee_applied = set()
        for t in temp:
            d = t["driver"]
            iban = t["iban"]
            group_key = (t["category"], iban)
            fee = 0.0
            if group_key in fee_groups and group_key not in fee_applied:
                fee = 1.25
                fee_applied.add(group_key)
            group_total = iban_groups.get(group_key, t["net"] - fee)
            # Prefixo visível usado também para separar os XMLs.
            origins_with_category = f"{t['category']} | {t['origins']}"
            items.append((
                closing_id, int(d["id"]), str(d.get("name") or "SEM NOME"), iban,
                str(d.get("bank_color") or ""), origins_with_category,
                t["gross"], t["cash"], t["commission"], t["fuel"],
                t["discount"], t["reimbursement"], t["immediate"], fee,
                t["net"], group_total,
            ))

        status = f"PROCESSADO ({unmatched} não encontrados; {invalid_rows} linhas ignoradas)"
        with db() as con:
            con.execute("DELETE FROM closing_items WHERE closing_id=?", (closing_id,))
            if items:
                con.executemany(
                    """INSERT INTO closing_items(
                    closing_id,driver_id,driver_name,iban,bank_color,origins,gross,cash,
                    commission,fuel,discount,reimbursement,immediate,bank_fee,
                    net_before_group,group_total)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    items,
                )
            con.execute("UPDATE closings SET status=? WHERE id=?", (status, closing_id))
        return closing_id
    except Exception as exc:
        app.logger.exception("Falha ao gerar fechamento %s", closing_id)
        with db() as con:
            con.execute(
                "UPDATE closings SET status=? WHERE id=?",
                (f"ERRO: {type(exc).__name__}: {str(exc)[:300]}", closing_id),
            )
        raise

def _payment_rows(closing_id: int, category: str):
    prefix = f"{category} |%"
    with db() as con:
        return con.execute(
            """SELECT iban, MAX(driver_name) driver_name, MAX(group_total) amount
            FROM closing_items
            WHERE closing_id=? AND iban<>'' AND origins LIKE ?
            GROUP BY iban
            HAVING amount>0
            ORDER BY driver_name""",
            (closing_id, prefix),
        ).fetchall()


def _split_payment_rows(rows, limit: float = 50000.0):
    parts, current, current_total = [], [], 0.0
    for row in rows:
        amount = money(row["amount"])
        if current and current_total + amount > limit:
            parts.append(current)
            current, current_total = [], 0.0
        current.append(row)
        current_total += amount
    if current:
        parts.append(current)
    return parts


def generate_xml_part(closing_id: int, rows, part_number: int = 1) -> bytes:
    ns = "urn:iso:std:iso:20022:tech:xsd:pain.001.001.03"
    doc = Element("Document", xmlns=ns)
    init = SubElement(doc, "CstmrCdtTrfInitn")
    grp = SubElement(init, "GrpHdr")
    msg_id = f"FLEET-{closing_id}-P{part_number:02d}-{datetime.now():%Y%m%d%H%M%S}"
    SubElement(grp, "MsgId").text = msg_id
    SubElement(grp, "CreDtTm").text = datetime.now().isoformat(timespec="seconds")
    SubElement(grp, "NbOfTxs").text = str(len(rows))
    total = sum(money(r["amount"]) for r in rows)
    SubElement(grp, "CtrlSum").text = f"{total:.2f}"
    initg = SubElement(grp, "InitgPty")
    SubElement(initg, "Nm").text = COMPANY_NAME
    pmt = SubElement(init, "PmtInf")
    SubElement(pmt, "PmtInfId").text = msg_id
    SubElement(pmt, "PmtMtd").text = "TRF"
    SubElement(pmt, "BtchBookg").text = "true"
    SubElement(pmt, "NbOfTxs").text = str(len(rows))
    SubElement(pmt, "CtrlSum").text = f"{total:.2f}"
    pmt_type = SubElement(pmt, "PmtTpInf")
    svc = SubElement(pmt_type, "SvcLvl")
    SubElement(svc, "Cd").text = "SEPA"
    SubElement(pmt, "ReqdExctnDt").text = datetime.now().date().isoformat()
    dbtr = SubElement(pmt, "Dbtr")
    SubElement(dbtr, "Nm").text = COMPANY_NAME
    dbtr_acct = SubElement(pmt, "DbtrAcct")
    dbtr_id = SubElement(dbtr_acct, "Id")
    SubElement(dbtr_id, "IBAN").text = os.getenv("DEBTOR_IBAN", "PT50000000000000000000000")
    dbtr_agt = SubElement(pmt, "DbtrAgt")
    fin = SubElement(dbtr_agt, "FinInstnId")
    SubElement(fin, "BIC").text = os.getenv("DEBTOR_BIC", "NOTPROVIDED")
    SubElement(pmt, "ChrgBr").text = "SLEV"
    for i, r in enumerate(rows, 1):
        tx = SubElement(pmt, "CdtTrfTxInf")
        pid = SubElement(tx, "PmtId")
        SubElement(pid, "EndToEndId").text = f"{msg_id}-{i}"
        amt = SubElement(tx, "Amt")
        SubElement(amt, "InstdAmt", Ccy="EUR").text = f"{money(r['amount']):.2f}"
        cdtr_agt = SubElement(tx, "CdtrAgt")
        SubElement(SubElement(cdtr_agt, "FinInstnId"), "BIC").text = "NOTPROVIDED"
        cdtr = SubElement(tx, "Cdtr")
        SubElement(cdtr, "Nm").text = str(r["driver_name"])[:70]
        aid = SubElement(SubElement(tx, "CdtrAcct"), "Id")
        SubElement(aid, "IBAN").text = r["iban"]
        SubElement(SubElement(tx, "RmtInf"), "Ustrd").text = f"Fechamento {closing_id}"
    return tostring(doc, encoding="utf-8", xml_declaration=True)


@app.context_processor
def inject_globals():
    return {"company_name": COMPANY_NAME, "app_version": APP_VERSION}


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("username") == ADMIN_USER and request.form.get("password") == ADMIN_PASSWORD:
            session["user"] = ADMIN_USER
            return redirect(url_for("dashboard"))
        flash("Utilizador ou senha incorretos.", "danger")
    return render_template("login.html")


@app.route("/company-logo")
def company_logo():
    """Entrega a logo usada no login e no menu sem depender da pasta static."""
    logo_path = BASE_DIR / "logo_empresa.png"
    if not logo_path.exists():
        # Fallback transparente para nunca derrubar o login por falta do ficheiro.
        from flask import Response
        import base64
        pixel = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII=")
        return Response(pixel, mimetype="image/png")
    return send_file(logo_path, mimetype="image/png", max_age=3600)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    with db() as con:
        drivers = con.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
        imports = con.execute("SELECT COUNT(*) FROM imports WHERE status='OK'").fetchone()[0]
        last = con.execute("SELECT * FROM closings ORDER BY id DESC LIMIT 1").fetchone()
        stats = {"gross":0,"commission":0,"fuel":0,"discount":0,"reimbursement":0,"net":0,"payments":0,"missing_iban":0}
        origins = []
        if last:
            # A Dashboard não recalcula dinheiro em mãos, descontos, combustível
            # ou reembolsos. Ela lê diretamente as mesmas linhas utilizadas no
            # relatório geral do fechamento, garantindo totais idênticos.
            r = con.execute("""SELECT
                    COALESCE(SUM(gross),0) gross,
                    COALESCE(SUM(fuel),0) fuel,
                    COALESCE(SUM(discount),0) discount,
                    COALESCE(SUM(reimbursement),0) reimbursement,
                    COALESCE(SUM(cash),0) cash,
                    COALESCE(SUM(net_before_group-bank_fee),0) net,
                    COUNT(DISTINCT CASE WHEN TRIM(COALESCE(iban,''))<>'' THEN iban END) payments,
                    COUNT(DISTINCT CASE WHEN TRIM(COALESCE(iban,''))='' AND net_before_group>0 THEN driver_id END) missing_iban
                FROM closing_items WHERE closing_id=?""", (last["id"],)).fetchone()
            stats = dict(r)
            shares = con.execute("""WITH base AS (
                    SELECT driver_id, CASE WHEN origins LIKE 'TVDE |%' THEN 'TVDE' ELSE 'DELIVERY' END AS category, MAX(commission) commission
                    FROM closing_items WHERE closing_id=? GROUP BY driver_id, CASE WHEN origins LIKE 'TVDE |%' THEN 'TVDE' ELSE 'DELIVERY' END
                )
                SELECT
                    COALESCE(SUM(base.commission * CASE
                        WHEN COALESCE(d.commission_owner,0)>1 THEN d.commission_owner/100.0
                        ELSE COALESCE(d.commission_owner,0) END),0) owner_total,
                    COALESCE(SUM(base.commission * CASE
                        WHEN COALESCE(d.partner_commission,0)>1 THEN d.partner_commission/100.0
                        ELSE COALESCE(d.partner_commission,0) END),0) partner_total
                FROM base JOIN drivers d ON d.id=base.driver_id""", (last["id"],)).fetchone()
            stats["commission"] = float(shares["owner_total"] or 0)
            stats["partner_commission"] = float(shares["partner_total"] or 0)
            origins = con.execute("SELECT origins,SUM(gross) total FROM closing_items WHERE closing_id=? GROUP BY origins ORDER BY total DESC", (last["id"],)).fetchall()
            fuel_breakdown = con.execute("""SELECT source_file, COALESCE(SUM(amount),0) total
                FROM fuel GROUP BY source_file ORDER BY total DESC""").fetchall()
        else:
            fuel_breakdown = []
        recent = con.execute("SELECT * FROM closings ORDER BY id DESC LIMIT 8").fetchall()
    origin_labels = [row["origins"] or "Sem origem" for row in origins]
    origin_values = [float(row["total"] or 0) for row in origins]
    return render_template("dashboard.html", drivers=drivers, imports=imports, last=last, stats=stats, origins=origins, recent=recent, origin_labels=origin_labels, origin_values=origin_values, fuel_breakdown=fuel_breakdown)


@app.route("/drivers")
@login_required
def drivers():
    q = request.args.get("q", "").strip()
    with db() as con:
        if q:
            like = f"%{q}%"
            rows = con.execute("SELECT * FROM drivers WHERE name LIKE ? OR external_id LIKE ? OR iban LIKE ? OR city LIKE ? OR partner LIKE ? ORDER BY name LIMIT 500", (like,like,like,like,like)).fetchall()
        else:
            rows = con.execute("SELECT * FROM drivers ORDER BY name LIMIT 500").fetchall()
        total = con.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
    return render_template("drivers.html", rows=rows, q=q, total=total)


@app.route("/drivers/<int:driver_id>", methods=["GET", "POST"])
@login_required
def edit_driver(driver_id: int):
    with db() as con:
        d = con.execute("SELECT * FROM drivers WHERE id=?", (driver_id,)).fetchone()
        if not d:
            return "Motorista não encontrado", 404
        if request.method == "POST":
            fields = ["external_id","name","city","company","iban","bank_color","partner","fuel_card","rental_label","observation"]
            vals = [request.form.get(f, "") for f in fields]
            nums = [money(request.form.get(f, 0)) for f in ("commission_owner","partner_commission","percentage","discount","reimbursement","immediate")]
            con.execute("""UPDATE drivers SET external_id=?,name=?,city=?,company=?,iban=?,bank_color=?,partner=?,fuel_card=?,rental_label=?,observation=?,commission_owner=?,partner_commission=?,percentage=?,discount=?,reimbursement=?,immediate=?,updated_at=? WHERE id=?""", (*vals,*nums,datetime.now().isoformat(timespec="seconds"),driver_id))
            flash("Cadastro atualizado.", "success")
            return redirect(url_for("edit_driver", driver_id=driver_id))
    return render_template("driver_edit.html", d=d)


@app.route("/database", methods=["GET", "POST"])
@login_required
def database_page():
    if request.method == "POST":
        f = request.files.get("database_file")
        if not f or not f.filename.lower().endswith(".xlsx"):
            flash("Envie um ficheiro XLSX válido.", "danger")
        else:
            p = UPLOAD_DIR / secure_filename(f.filename)
            f.save(p)
            try:
                n, m = import_master_workbook(p)
                flash(f"Banco atualizado: {n} motoristas e {m} configurações de leitura.", "success")
            except Exception as exc:
                flash(f"Erro ao importar banco: {exc}", "danger")
    with db() as con:
        n = con.execute("SELECT COUNT(*) FROM drivers").fetchone()[0]
        m = con.execute("SELECT COUNT(*) FROM mappings").fetchone()[0]
    return render_template("database.html", n=n, m=m)


@app.route("/imports", methods=["GET", "POST"])
@login_required
def imports_page():
    if request.method == "POST":
        kind = request.form.get("kind", "reports")
        files = request.files.getlist("files")
        success = errors = 0
        for f in files:
            if not f or not f.filename:
                continue
            filename = secure_filename(f.filename)
            data = f.read()
            # A coluna sha256 possui UNIQUE no banco das versões anteriores.
            # Para permitir arquivos de empresas diferentes com conteúdo idêntico,
            # a assinatura considera o nome normalizado + o conteúdo do arquivo.
            # Assim, somente o mesmo arquivo (mesmo nome e conteúdo) é duplicado.
            signature_source = filename.casefold().encode("utf-8") + b"\0" + data
            sha = hashlib.sha256(signature_source).hexdigest()
            with db() as con:
                duplicate = con.execute(
                    "SELECT id FROM imports WHERE sha256=?",
                    (sha,),
                ).fetchone()
            if duplicate:
                flash(f"{filename}: este mesmo arquivo já foi importado nesta semana; ignorado.", "warning")
                continue
            p = UPLOAD_DIR / f"{datetime.now():%Y%m%d%H%M%S%f}_{filename}"
            p.write_bytes(data)
            try:
                with db() as con:
                    cur = con.execute(
                        "INSERT INTO imports(filename,sha256,kind,status,created_at) VALUES(?,?,?,?,?)",
                        (filename, sha, kind, "PROCESSANDO", datetime.now().isoformat(timespec="seconds")),
                    )
                    iid = cur.lastrowid
            except sqlite3.IntegrityError:
                # Proteção adicional para bases antigas ou envios simultâneos.
                flash(f"{filename}: este mesmo arquivo já foi importado nesta semana; ignorado.", "warning")
                continue
            try:
                if kind == "fuel" or filename.lower().endswith((".xlsx", ".xls")):
                    count, msg = process_fuel_file(p, iid)
                else:
                    count, msg = process_report_file(p, iid)
                with db() as con:
                    con.execute("UPDATE imports SET status='OK',rows_count=?,message=? WHERE id=?", (count,msg,iid))
                success += 1
            except Exception as exc:
                with db() as con:
                    con.execute("UPDATE imports SET status='ERRO',message=? WHERE id=?", (str(exc),iid))
                errors += 1
        flash(f"Importação concluída: {success} ficheiros processados e {errors} com erro.", "success" if errors == 0 else "warning")
        return redirect(url_for("imports_page"))
    with db() as con:
        rows = con.execute("SELECT * FROM imports ORDER BY id DESC LIMIT 100").fetchall()
    return render_template("imports.html", rows=rows)


@app.route("/imports/clear", methods=["POST"])
@login_required
def clear_imports():
    with db() as con:
        con.execute("DELETE FROM raw_earnings")
        con.execute("DELETE FROM fuel")
        con.execute("DELETE FROM imports")
    flash("Importações da semana foram limpas.", "success")
    return redirect(url_for("imports_page"))


@app.route("/closings", methods=["GET", "POST"])
@login_required
def closings_page():
    if request.method == "POST":
        label = request.form.get("label") or f"Semana {datetime.now():%d/%m/%Y}"
        try:
            cid = build_closing(label)
            flash("Fechamento processado. Confira os valores antes de gerar o XML.", "success")
            return redirect(url_for("closing_detail", closing_id=cid))
        except Exception as exc:
            app.logger.exception("Erro no botão Processar semana")
            flash(f"Não foi possível gerar o fechamento: {type(exc).__name__}: {str(exc)}", "danger")
            return redirect(url_for("closings_page"))
    with db() as con:
        rows = con.execute("SELECT * FROM closings ORDER BY id DESC").fetchall()
    return render_template("closings.html", rows=rows)


@app.route("/closings/<int:closing_id>")
@login_required
def closing_detail(closing_id: int):
    with db() as con:
        closing = con.execute("SELECT * FROM closings WHERE id=?", (closing_id,)).fetchone()
        items = con.execute("SELECT * FROM closing_items WHERE closing_id=? ORDER BY driver_name", (closing_id,)).fetchall()
    return render_template("closing_detail.html", closing=closing, items=items)


@app.route("/closings/<int:closing_id>/excel")
@login_required
def closing_excel(closing_id: int):
    with db() as con:
        df = pd.read_sql_query("SELECT driver_name AS Motorista,iban AS IBAN,bank_color AS Banco,origins AS Origens,gross AS Bruto,cash AS Dinheiro,commission AS Comissao,fuel AS Combustivel,discount AS Desconto,reimbursement AS Reembolso,immediate AS Imediata,bank_fee AS Taxa_Banco,net_before_group AS Liquido_Individual,group_total AS Total_IBAN FROM closing_items WHERE closing_id=? ORDER BY driver_name", con, params=(closing_id,))
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resultado_Geral")
        same = df[df["IBAN"].duplicated(False) & (df["IBAN"] != "")]
        same.to_excel(writer, index=False, sheet_name="Mesmo_IBAN")
        df[df["IBAN"] == ""].to_excel(writer, index=False, sheet_name="Sem_IBAN")
        df[df["Liquido_Individual"] < 0].to_excel(writer, index=False, sheet_name="Negativos")
        rentals = df[df["Desconto"] > 0][["Motorista","Desconto","Origens"]]
        rentals.to_excel(writer, index=False, sheet_name="Descontos_Alugueis")
        style_workbook(writer.book)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name=f"fechamento_{closing_id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/closings/<int:closing_id>/xml")
@login_required
def closing_xml(closing_id: int):
    grouped_parts = []
    for category, filename_label in (("TVDE", "PAGAMENTO_TVDE"), ("DELIVERY", "PAGAMENTO_DELIVERY")):
        rows = _payment_rows(closing_id, category)
        parts = _split_payment_rows(rows, 50000.0)
        for index, part in enumerate(parts, 1):
            grouped_parts.append((category, filename_label, index, part))

    if not grouped_parts:
        flash("Não existem pagamentos positivos com IBAN para gerar XML.", "warning")
        return redirect(url_for("closing_detail", closing_id=closing_id))

    # Entrega sempre um ZIP para manter TVDE e Delivery claramente separados.
    out = io.BytesIO()
    counters = {"TVDE": 0, "DELIVERY": 0}
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for category, filename_label, index, part in grouped_parts:
            counters[category] += 1
            xml = generate_xml_part(closing_id, part, index)
            zf.writestr(f"{filename_label}_{index:02d}.xml", xml)
    out.seek(0)
    total_files = len(grouped_parts)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"pagamentos_{closing_id}_TVDE_DELIVERY_{total_files}_xml.zip",
        mimetype="application/zip",
    )




def refresh_closing_payment_groups(closing_id: int) -> None:
    """Recalcula apenas consolidação por IBAN e taxa bancária, sem alterar os cálculos financeiros."""
    with db() as con:
        rows = con.execute("SELECT * FROM closing_items WHERE closing_id=? ORDER BY id", (closing_id,)).fetchall()
        groups = defaultdict(list)
        for row in rows:
            category = (row["origins"] or "").split(" |", 1)[0].strip() or "OUTROS"
            iban = (row["iban"] or "").strip().upper().replace(" ", "")
            if iban:
                groups[(category, iban)].append(row)
        con.execute("UPDATE closing_items SET bank_fee=0, group_total=net_before_group WHERE closing_id=?", (closing_id,))
        for (_category, iban), items in groups.items():
            total = sum(money(r["net_before_group"]) for r in items)
            charge_fee = any(norm(r["bank_color"]) == "AZUL" for r in items)
            if charge_fee:
                total -= 1.25
            first_id = items[0]["id"]
            for r in items:
                con.execute("UPDATE closing_items SET group_total=?, bank_fee=? WHERE id=?", (total, 1.25 if charge_fee and r["id"] == first_id else 0.0, r["id"]))

MISSING_IBAN_SQL = "UPPER(TRIM(COALESCE({col},''))) IN ('', '-', '0', '0.0', 'NAN', 'NONE', 'NULL', 'N/A', 'NA', 'SEM IBAN')"


@app.route("/missing-iban")
@login_required
def missing_iban_page():
    q = request.args.get("q", "").strip()
    missing_values = "('', '-', '0', '0.0', 'NAN', 'NONE', 'NULL', 'N/A', 'NA', 'SEM IBAN')"
    with db() as con:
        last = con.execute("SELECT id FROM closings WHERE status LIKE 'PROCESSADO%' ORDER BY id DESC LIMIT 1").fetchone()
        params = []
        pending_join = "LEFT JOIN (SELECT NULL driver_id, 0 pending_net WHERE 0) p ON p.driver_id=d.id"
        if last:
            pending_join = f"""LEFT JOIN (
                SELECT driver_id, SUM(net_before_group-bank_fee) pending_net
                FROM closing_items
                WHERE closing_id=? AND UPPER(TRIM(COALESCE(iban,''))) IN {missing_values}
                GROUP BY driver_id
            ) p ON p.driver_id=d.id"""
            params.append(last["id"])
        search_sql = ""
        if q:
            like = f"%{q}%"
            search_sql = " AND (d.name LIKE ? OR d.external_id LIKE ? OR d.partner LIKE ? OR d.city LIKE ?)"
            params.extend([like, like, like, like])
        # A página mostra TODOS os cadastros sem IBAN, mesmo quando ainda não
        # existe fechamento ou quando o líquido pendente é zero/negativo.
        # O fechamento serve apenas para exibir o valor pendente ao lado.
        rows = con.execute(f"""
            SELECT d.*, COALESCE(p.pending_net,0) pending_net
            FROM drivers d
            {pending_join}
            WHERE UPPER(TRIM(COALESCE(d.iban,''))) IN {missing_values}
            {search_sql}
            ORDER BY CASE WHEN COALESCE(p.pending_net,0)>0 THEN 0 ELSE 1 END, d.name
            LIMIT 1000
        """, params).fetchall()
        total_missing = con.execute(f"SELECT COUNT(*) FROM drivers WHERE UPPER(TRIM(COALESCE(iban,''))) IN {missing_values}").fetchone()[0]
        closing_missing = 0
        if last:
            closing_missing = con.execute(f"""SELECT COUNT(*) FROM (
                SELECT driver_id
                FROM closing_items
                WHERE closing_id=? AND UPPER(TRIM(COALESCE(iban,''))) IN {missing_values}
                GROUP BY driver_id
                HAVING SUM(net_before_group-bank_fee)>0
            )""", (last["id"],)).fetchone()[0]
    return render_template("missing_iban.html", rows=rows, q=q, total_missing=total_missing, closing_missing=closing_missing)


@app.route("/missing-iban/<int:driver_id>/save", methods=["POST"])
@login_required
def save_missing_iban(driver_id: int):
    iban = re.sub(r"\s+", "", request.form.get("iban", "").upper())
    if len(iban) < 15:
        flash("Informe um IBAN válido antes de salvar.", "danger")
        return redirect(url_for("missing_iban_page"))
    bank_color = norm(request.form.get("bank_color", "YELLOW"))
    if bank_color not in {"AZUL", "YELLOW"}:
        bank_color = "YELLOW"
    partner = request.form.get("partner", "").strip()
    partner_commission = money(request.form.get("partner_commission", 0))
    with db() as con:
        driver = con.execute("SELECT * FROM drivers WHERE id=?", (driver_id,)).fetchone()
        if not driver:
            return "Motorista não encontrado", 404
        # Atualiza pelo identificador da plataforma, nunca apenas pelo nome.
        # Nomes duplicados permanecem independentes; linhas realmente duplicadas
        # com o mesmo ID externo recebem a correção em conjunto.
        external_key = clean_identifier(driver["external_id"])
        if external_key:
            duplicate_rows = con.execute("SELECT id,external_id FROM drivers").fetchall()
            related_ids = [int(r["id"]) for r in duplicate_rows if clean_identifier(r["external_id"]) == external_key]
        else:
            related_ids = [driver_id]
        if not related_ids:
            related_ids = [driver_id]
        placeholders = ",".join("?" for _ in related_ids)
        now = datetime.now().isoformat(timespec="seconds")
        con.execute(f"UPDATE drivers SET iban=?,bank_color=?,partner=?,partner_commission=?,updated_at=? WHERE id IN ({placeholders})",
                    [iban, bank_color, partner, partner_commission, now, *related_ids])
        closings = con.execute(f"SELECT DISTINCT closing_id FROM closing_items WHERE driver_id IN ({placeholders}) AND UPPER(TRIM(COALESCE(iban,''))) IN ('', '-', '0', '0.0', 'NAN', 'NONE', 'NULL', 'N/A', 'NA', 'SEM IBAN')", related_ids).fetchall()
        con.execute(f"UPDATE closing_items SET iban=?,bank_color=? WHERE driver_id IN ({placeholders}) AND UPPER(TRIM(COALESCE(iban,''))) IN ('', '-', '0', '0.0', 'NAN', 'NONE', 'NULL', 'N/A', 'NA', 'SEM IBAN')",
                    [iban, bank_color, *related_ids])
    for row in closings:
        refresh_closing_payment_groups(row["closing_id"])
    flash(f"IBAN de {driver['name']} atualizado em todos os cadastros e fechamento recalculado.", "success")
    return redirect(url_for("missing_iban_page"))


@app.route("/missing-iban/excel")
@login_required
def missing_iban_excel():
    with db() as con:
        df = pd.read_sql_query("SELECT name AS Motorista,external_id AS ID,city AS Cidade,company AS Companhia,partner AS Parceiro,partner_commission AS Comissao_Parceiro,bank_color AS Banco,iban AS IBAN FROM drivers WHERE UPPER(TRIM(COALESCE(iban,''))) IN ('', '-', '0', '0.0', 'NAN', 'NONE', 'NULL', 'N/A', 'NA', 'SEM IBAN') ORDER BY name", con)
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sem_IBAN")
        if not df.empty:
            for partner, group in df.groupby(df["Parceiro"].fillna("").replace("", "SEM_PARCEIRO")):
                safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(partner))[:25] or "SEM_PARCEIRO"
                group.to_excel(writer, index=False, sheet_name=safe[:31])
        style_workbook(writer.book)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="motoristas_sem_iban.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/partners")
@login_required
def partners_page():
    with db() as con:
        last = con.execute("SELECT id FROM closings WHERE status LIKE 'PROCESSADO%' ORDER BY id DESC LIMIT 1").fetchone()
        if not last:
            return render_template("partners.html", rows=[], closing_id=0)
        rows = con.execute(
            """SELECT d.partner AS partner, COUNT(DISTINCT ci.driver_id) AS drivers,
            COALESCE(SUM(ci.gross),0) AS gross, COALESCE(SUM(ci.net_before_group-ci.bank_fee),0) AS net
            FROM closing_items ci JOIN drivers d ON d.id=ci.driver_id
            WHERE ci.closing_id=? AND TRIM(COALESCE(d.partner,''))<>''
            GROUP BY d.partner ORDER BY d.partner""",
            (last["id"],),
        ).fetchall()
    return render_template("partners.html", rows=rows, closing_id=last["id"])


@app.route("/partners/<int:closing_id>/<path:partner>/excel")
@login_required
def partner_excel(closing_id: int, partner: str):
    with db() as con:
        df = pd.read_sql_query(
            """SELECT d.partner AS Parceiro, ci.driver_name AS Motorista, ci.origins AS Origens,
            ci.gross AS Bruto, ci.cash AS Dinheiro, ci.commission AS Comissao, ci.fuel AS Combustivel,
            ci.discount AS Desconto, ci.reimbursement AS Reembolso,
            d.partner_commission AS Comissao_Parceiro,
            ci.net_before_group-ci.bank_fee AS Liquido, ci.iban AS IBAN
            FROM closing_items ci JOIN drivers d ON d.id=ci.driver_id
            WHERE ci.closing_id=? AND d.partner=? ORDER BY ci.driver_name""",
            con, params=(closing_id, partner),
        )
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Motoristas")
        pd.DataFrame([{
            "Parceiro": partner,
            "Motoristas": int(df["Motorista"].nunique()) if not df.empty else 0,
            "Total_Bruto": float(df["Bruto"].sum()) if not df.empty else 0,
            "Total_Comissao_Parceiro": float(df["Comissao_Parceiro"].sum()) if not df.empty else 0,
            "Total_Liquido": float(df["Liquido"].sum()) if not df.empty else 0,
        }]).to_excel(writer, index=False, sheet_name="Resumo")
        style_workbook(writer.book)
    out.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", partner).strip("_") or "parceiro"
    return send_file(out, as_attachment=True, download_name=f"parceiro_{safe}_fechamento_{closing_id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



def _pt_money(value: Any) -> str:
    rendered = f"{float(value or 0):,.2f}"
    return "€ " + rendered.replace(",", "X").replace(".", ",").replace("X", ".")


@app.route("/fechamentos/<int:closing_id>/recibo/<int:driver_id>")
@login_required
def driver_receipt(closing_id: int, driver_id: int):
    with db() as con:
        closing = con.execute("SELECT * FROM closings WHERE id=?", (closing_id,)).fetchone()
        driver = con.execute("SELECT * FROM drivers WHERE id=?", (driver_id,)).fetchone()
        rows = con.execute("""SELECT * FROM closing_items
            WHERE closing_id=? AND driver_id=? ORDER BY origins""", (closing_id, driver_id)).fetchall()
    if not closing or not driver or not rows:
        return "Pagamento não encontrado", 404

    totals = {key: sum(float(r[key] or 0) for r in rows) for key in
              ("gross", "cash", "commission", "fuel", "discount", "reimbursement", "immediate", "bank_fee")}
    totals["net"] = sum(float((r["net_before_group"] or 0) - (r["bank_fee"] or 0)) for r in rows)
    origins = " / ".join(dict.fromkeys(str(r["origins"] or "") for r in rows))

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm,
                            topMargin=16*mm, bottomMargin=16*mm, title=f"Recibo - {driver['name']}")
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReceiptTitle", parent=styles["Title"], fontSize=18, leading=22, alignment=TA_CENTER, textColor=colors.HexColor("#17324D"))
    small_center = ParagraphStyle("SmallCenter", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor("#667085"))
    right_style = ParagraphStyle("Right", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=10)
    story = []
    logo = BASE_DIR / "logo_empresa.png"
    if logo.exists():
        try:
            story.append(Image(str(logo), width=42*mm, height=24*mm, kind="proportional"))
        except Exception:
            pass
    story += [Paragraph(COMPANY_NAME, title_style), Paragraph("RECIBO DE PAGAMENTO DO MOTORISTA", small_center), Spacer(1, 7*mm)]
    info = [
        ["Motorista", str(driver["name"] or "")],
        ["ID / Identificador", str(driver["external_id"] or "")],
        ["Período / Fechamento", str(closing["label"] or "")],
        ["Origem", origins],
        ["IBAN", str(rows[0]["iban"] or "SEM IBAN")],
        ["Data de emissão", datetime.now().strftime("%d/%m/%Y %H:%M")],
    ]
    info_table = Table(info, colWidths=[48*mm, 112*mm])
    info_table.setStyle(TableStyle([("BACKGROUND",(0,0),(0,-1),colors.HexColor("#EAF2FF")),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D0D5DD")),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),7)]))
    story += [info_table, Spacer(1, 7*mm)]
    financial = [
        ["Descrição", "Valor"],
        ["Valor bruto", _pt_money(totals["gross"])],
        ["Dinheiro em mãos", _pt_money(totals["cash"])],
        ["Comissão", "- " + _pt_money(totals["commission"])],
        ["Combustível", "- " + _pt_money(totals["fuel"])],
        ["Descontos", "- " + _pt_money(totals["discount"])],
        ["Pagamento imediato", "- " + _pt_money(totals["immediate"])],
        ["Reembolso", "+ " + _pt_money(totals["reimbursement"])],
        ["Taxa bancária", "- " + _pt_money(totals["bank_fee"])],
        ["VALOR LÍQUIDO PAGO", _pt_money(totals["net"])],
    ]
    fin_table = Table(financial, colWidths=[105*mm, 55*mm])
    fin_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#17324D")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("ALIGN",(1,1),(1,-1),"RIGHT"),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D0D5DD")),("PADDING",(0,0),(-1,-1),8),("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#DCFCE7")),("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,-1),(-1,-1),12)]))
    story += [fin_table, Spacer(1, 14*mm), Paragraph("Este recibo foi gerado automaticamente com base no fechamento registado na plataforma.", small_center)]
    doc.build(story)
    out.seek(0)
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(driver["name"] or "motorista")).strip("_")
    return send_file(out, as_attachment=True, download_name=f"recibo_{safe}_fechamento_{closing_id}.pdf", mimetype="application/pdf")


@app.route("/gestor")
@login_required
def manager_report():
    try:
        with db() as con:
            closing = con.execute("SELECT * FROM closings WHERE status LIKE 'PROCESSADO%' ORDER BY id DESC LIMIT 1").fetchone()
            if not closing:
                return render_template("manager.html", closing=None, stats={}, partners=[], negatives=[], missing=[], fuel_breakdown=[])
            cid = closing["id"]
            items = con.execute("""SELECT ci.*, COALESCE(d.partner,'') partner,
                COALESCE(d.commission_owner,0) owner_share, COALESCE(d.partner_commission,0) partner_share
                FROM closing_items ci LEFT JOIN drivers d ON d.id=ci.driver_id WHERE ci.closing_id=?""", (cid,)).fetchall()
            stats={"gross":0.0,"net":0.0,"fuel":0.0,"discount":0.0,"reimbursement":0.0,"cash":0.0,"negatives":0,"missing_iban":0,"tvde":0.0,"delivery":0.0,"commission":0.0,"partner_commission":0.0}
            per_driver={}
            partners_map={}
            for r in items:
                stats["gross"] += float(r["gross"] or 0); stats["net"] += float((r["net_before_group"] or 0)-(r["bank_fee"] or 0))
                stats["fuel"] += float(r["fuel"] or 0); stats["discount"] += float(r["discount"] or 0); stats["reimbursement"] += float(r["reimbursement"] or 0); stats["cash"] += float(r["cash"] or 0)
                if str(r["origins"] or '').startswith('TVDE |'): stats["tvde"] += float(r["gross"] or 0)
                else: stats["delivery"] += float(r["gross"] or 0)
                owner=float(r["owner_share"] or 0); owner=owner/100 if owner>1 else owner
                pshare=float(r["partner_share"] or 0); pshare=pshare/100 if pshare>1 else pshare
                stats["commission"] += float(r["commission"] or 0)*owner
                stats["partner_commission"] += float(r["commission"] or 0)*pshare
                did=r["driver_id"]; d=per_driver.setdefault(did,{"driver_name":r["driver_name"],"origins":[],"partner":r["partner"],"net":0.0,"iban":r["iban"] or ''})
                d["origins"].append(r["origins"] or ''); d["net"] += float((r["net_before_group"] or 0)-(r["bank_fee"] or 0))
                pn=(r["partner"] or 'SEM PARCEIRO').strip() or 'SEM PARCEIRO'; pm=partners_map.setdefault(pn,{"partner":pn,"ids":set(),"gross":0.0,"partner_commission":0.0,"net":0.0})
                pm["ids"].add(did); pm["gross"] += float(r["gross"] or 0); pm["partner_commission"] += float(r["commission"] or 0)*pshare; pm["net"] += float((r["net_before_group"] or 0)-(r["bank_fee"] or 0))
            negatives=[]; missing=[]
            for d in per_driver.values():
                row={"driver_name":d["driver_name"],"origins":" / ".join(dict.fromkeys(d["origins"])),"partner":d["partner"],"net":d["net"]}
                if d["net"]<0: negatives.append(row)
                if not str(d["iban"]).strip(): missing.append(row)
            stats["negatives"]=len(negatives); stats["missing_iban"]=len(missing)
            partners=[]
            for pm in partners_map.values():
                partners.append({"partner":pm["partner"],"drivers":len(pm["ids"]),"gross":pm["gross"],"partner_commission":pm["partner_commission"],"net":pm["net"]})
            partners.sort(key=lambda x:x["gross"], reverse=True); negatives.sort(key=lambda x:x["net"]); missing.sort(key=lambda x:x["driver_name"] or '')
            fuel_breakdown=con.execute("SELECT source_file, COALESCE(SUM(amount),0) total FROM fuel GROUP BY source_file ORDER BY total DESC").fetchall()
        return render_template("manager.html",closing=closing,stats=stats,partners=partners,negatives=negatives,missing=missing,fuel_breakdown=fuel_breakdown)
    except Exception as exc:
        app.logger.exception("Erro na Área do Gestor")
        flash(f"Não foi possível abrir a Área do Gestor: {type(exc).__name__}", "danger")
        return redirect(url_for("dashboard"))

@app.route("/gestor/<int:closing_id>/excel")
@login_required
def manager_excel(closing_id:int):
    with db() as con:
        detail=pd.read_sql_query("""SELECT ci.driver_name AS Motorista,ci.origins AS Origem,d.city AS Cidade,d.partner AS Parceiro,
            ci.gross AS Bruto,ci.cash AS Dinheiro,ci.commission AS Comissao_Empresa,d.partner_commission AS Comissao_Parceiro,
            ci.fuel AS Combustivel,ci.discount AS Desconto,ci.reimbursement AS Reembolso,
            ci.net_before_group-ci.bank_fee AS Liquido,ci.iban AS IBAN FROM closing_items ci LEFT JOIN drivers d ON d.id=ci.driver_id
            WHERE ci.closing_id=? ORDER BY ci.driver_name""",con,params=(closing_id,))
    out=io.BytesIO()
    with pd.ExcelWriter(out,engine="openpyxl") as writer:
        detail.to_excel(writer,index=False,sheet_name="Geral")
        detail[detail["Liquido"]<0].to_excel(writer,index=False,sheet_name="Negativos")
        detail[detail["IBAN"].fillna("").str.strip()==""].to_excel(writer,index=False,sheet_name="Sem_IBAN")
        if not detail.empty:
            resumo=detail.groupby(detail["Parceiro"].fillna("").replace("","SEM PARCEIRO"),dropna=False).agg(Motoristas=("Motorista","nunique"),Bruto=("Bruto","sum"),Comissao_Parceiro=("Comissao_Parceiro","sum"),Liquido=("Liquido","sum")).reset_index()
            resumo.to_excel(writer,index=False,sheet_name="Parceiros")
        style_workbook(writer.book)
    out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"relatorio_gestor_{closing_id}.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/health")
def health():
    try:
        with db() as con:
            con.execute("SELECT 1").fetchone()
        return {"status": "ok", "database": str(DB_PATH)}, 200
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


@app.errorhandler(500)
def internal_error(exc):
    app.logger.error("Erro interno: %s\n%s", exc, traceback.format_exc())
    try:
        return render_template("error.html", message=f"Ocorreu um erro interno: {type(exc).__name__}. Consulte os logs do Render."), 500
    except Exception:
        return f"Erro interno: {type(exc).__name__}: {exc}", 500

@app.errorhandler(413)
def too_large(_):
    flash("O envio excede o limite de 80 MB.", "danger")
    return redirect(url_for("imports_page"))


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)
